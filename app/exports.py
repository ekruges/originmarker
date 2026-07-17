"""Exports for a pb.PanelResult: CSV, JSON, XLSX, PDF.

Every export is self-describing: build, both variant forms (genomic-VCF and transcript
sense), source versions, timestamps, the Layer-B protocol and pb.DISCLAIMER verbatim.
Heterozygosity columns are named "..._prior": 2pq is a population prior, never a claim
about this carrier.

    to_csv/to_json/to_xlsx/to_pdf(result) -> bytes
    FILENAME(result, ext) -> 'originmarker_rs151344623_GRCh38_2026-07-15.csv'
"""
from __future__ import annotations

import csv
import io
import json
import pathlib
import re
from xml.etree import ElementTree

import panelbuilder as pb
from app import ispcr

# gnomAD ancestry codes, in the engine's own order (AFR AMR ASJ EAS FIN NFE SAS MID).
POPS = list(pb.GNOMAD_POPS.values())

# The engine owns the star: the per-marker verdict, the rule and its words are all its to
# define. Exports read them and never recompute the predicate, so the page, the workbook
# and the web cannot drift into three different stars.
STAR_FIELD = pb.FLANKING_CRITERIA["field"]

# The primer pair as data, parallel to _primer_row(). Sequences and the verdict are columns
# and not prose, because the reader who strips the header block is the one who must not lose
# them. Every one is empty where the engine designed no primer for that marker.
PRIMER_COLUMNS = ("primer_fwd", "primer_fwd_tm", "primer_fwd_gc", "primer_fwd_len",
                  "primer_rev", "primer_rev_tm", "primer_rev_gc", "primer_rev_len",
                  "primer_product_bp", "primer_product_start", "primer_product_end",
                  "primer_masked_variants", "primer_insilico_pcr", "primer_warnings",
                  "primer_design_error")


def _columns(anc=None, primers=False) -> list:
    """Column names, parallel to _row(): keep the two in step, in order.

    Every quantity that decides in_recommended_panel must be a column, so the
    ancestry-matched 2pq joins the table whenever one was selected. The same rule already
    covers the star: every quantity STAR_FIELD is decided on is a column below.
    """
    return (["rsid", "chrom", "pos_grch38", "ref", "alt", "signed_dist_bp", "side", "tier",
             "maf", "af", "gnomad_an", "het_2pq_prior_global", "het_2pq_prior_max_pop"]
            + ([f"het_2pq_prior_{anc}"] if anc else [])
            + [f"maf_{p}" for p in POPS]
            + [f"an_{p}" for p in POPS]
            + ["cm_to_variant", "recomb_fraction", "hotspot_between", "map_approx",
               "ensembl_pos_check", "in_recommended_panel", STAR_FIELD]
            + (list(PRIMER_COLUMNS) if primers else []))


CSV_COLUMNS = tuple(_columns())            # the shape when no ancestry was selected


# --------------------------------------------------------------------------- #
# Shared facts. One source of truth for the header block / provenance sheet /
# PDF variant card, so the four formats cannot drift apart.
# --------------------------------------------------------------------------- #

def _ranking_key(result) -> str:
    """The engine's own name for the quantity that produced the sort order.

    Only panelbuilder knows the sort key, so exports render the name it gives and never
    restate the basis in their own words.
    """
    return result.provenance.get("ranking_key") or "not reported by this build"


def _rank_pop(result):
    """The ancestry the panel was ranked for, or None."""
    return result.provenance.get("ancestry_rank")


def _starred(m) -> bool:
    """Does this marker get the glyph? True only on the engine's own True: nothing here
    re-decides it, so an export cannot star a marker the engine did not."""
    return getattr(m, STAR_FIELD, None) is True


def _star_cell(m):
    """The verdict as DATA: True, False, or None where the engine never judged it.

    Three states, not two. The predicate runs over the shortlist and this column is written
    for every candidate, so flattening None to False would print a failing verdict on ~1200
    markers nobody assessed, several of them nearer the variant than the starred ones.
    """
    return getattr(m, STAR_FIELD, None)


def _flanking(result) -> dict:
    """The rule this panel was built under, as the engine stamped it into provenance.

    Read from the result, not from pb: the artifact must describe the rule that produced
    THIS panel, not whichever one the code holds when it is exported.
    """
    return result.provenance.get("flanking_criteria") or {}


def _star_legend(result, markers) -> list[str]:
    """The engine's own words for what the star claims, verbatim, or [] when no marker in
    markers carries one: a legend for an absent glyph teaches a symbol and then leaves the
    reader hunting the page for it.
    """
    if not any(_starred(m) for m in markers):
        return []
    return list(_flanking(result).get("note")
                or ["Star: the rule behind it is not reported by this build."])


def _star_count_text(result) -> str:
    """Per-side counts of markers meeting the criteria, read off the engine's coverage.

    Never recounted here: the count and the engine's own under-coverage flag have to be one
    verdict. ESHRE's minimum is per side, so a total would answer a question nobody asked.
    """
    cov, n = result.coverage, _flanking(result).get("min_per_side")
    counts = "; ".join(f"{cov[k]} {k.split('_')[0]}-coordinate"
                       for k in ("lower_flanking_count", "higher_flanking_count") if k in cov)
    return (f"{counts or 'not reported by this build'}. ESHRE recommends at least "
            f"{n if n is not None else 'three'} SNPs on each side of the pathogenic variant.")


def _star_facts(result) -> list[tuple[str, str]]:
    """The star's legend and per-side count, or nothing at all when nothing is starred."""
    note = _star_legend(result, result.candidates)
    return [] if not note else [(f"Star ({STAR_FIELD})", " ".join(note)),
                                ("Markers meeting the flanking criteria",
                                 _star_count_text(result))]


# --------------------------------------------------------------------------- #
# Primers. primers.py designs them and app/ispcr.py verifies them; both own every word
# of every warning they raise. Exports read the verdict and the words, and restate
# neither. Fields are read straight off primers.PrimerResult rather than defensively:
# a renamed field must break here loudly, not print an empty primer block.
# --------------------------------------------------------------------------- #

def _primer(m):
    """This marker's primers.PrimerResult, or None where the engine designed none.

    None means no design was attempted: primer3 is optional and absent by default. A design
    that ran and FAILED is a PrimerResult carrying `error`. Absence and failure are
    different facts and must never print the same.
    """
    return getattr(m, "primer", None)


def _has_primers(result) -> bool:
    return any(_primer(m) is not None for m in result.candidates)


def _pcr_state(p) -> str:
    """'danger' | 'pass' | 'unverified', from the engine's own verdict token.

    ispcr owns the vocabulary, so the tokens are read from it and never spelled here. Only
    ONE_PRODUCT is a pass: not_checked, unknown and any token a later build adds are all
    unverified, since a verdict this file failed to recognise must never render clean.
    """
    if p.insilico_pcr == ispcr.DANGER:
        return "danger"
    return "pass" if p.insilico_pcr == ispcr.ONE_PRODUCT else "unverified"


def _primer_warnings(p) -> list:
    """The engine's own warnings, verbatim and all of them.

    The LONG form of each: an export is read away from the app, so the short form's implicit
    "see the docs" would point at nothing. Never empty for a pair: primers.py welds them onto
    the result, and a primer printed with no warning beside it reads as a primer with nothing
    wrong with it.
    """
    return [x.long for x in p.warnings] or \
        ["What is wrong with this pair is not reported by this build. It has not been "
         "verified."]


def _mask_note(p) -> str:
    """The engine's own words for what the primer sites were kept clear of, in full.

    Never restated here: only the engine knows what its mask actually reached, and a label
    claiming more than was masked is worse than no label at all. The long form, like the
    warnings: an export is read away from the docs the short form defers to.
    """
    return p.mask_note.long if p.mask_note else \
        (f"{len(p.masked)} variants masked; what the mask covers is not reported by this "
         f"build.")


def _oligo_cells(o) -> list:
    """One oligo's four cells: sequence 5'->3' as ordered, Tm, GC, length."""
    return [None] * 4 if o is None else [o.seq, o.tm, o.gc, o.length]


def _primer_row(m) -> list:
    """Values in PRIMER_COLUMNS order. Keep the two in step."""
    p = _primer(m)
    if p is None:
        return [None] * len(PRIMER_COLUMNS)
    return [*_oligo_cells(p.fwd), *_oligo_cells(p.rev),
            # product_end is the engine's own property: never recomputed from start+size.
            p.product_size, p.product_start, p.product_end, len(p.masked),
            # The token verbatim: 'not_checked' is the engine's default and is not a pass.
            p.insilico_pcr, " | ".join(_primer_warnings(p)), p.error]


def _primer_dict(m):
    """The pair as JSON data, or None. Built here rather than left to asdict(): the warnings
    ride along as a list, and JSON must not be the one format that drops one.
    """
    p = _primer(m)
    if p is None:
        return None
    return dict(zip(PRIMER_COLUMNS, _primer_row(m)),
                primer_warnings=_primer_warnings(p), mask_note=_mask_note(p))


def _primer_facts(result) -> list[tuple[str, str]]:
    """Pair counts and what was checked, or nothing at all where none was designed: a
    "0 primers" line on a server that never ran a design reports a failure that never
    happened.
    """
    pm = [_primer(m) for m in result.candidates if _primer(m) is not None]
    if not pm:
        return []
    st = [_pcr_state(p) for p in pm]
    fail = sum(1 for p in pm if p.error)
    return [("Primer pairs", f"{len(pm) - fail} designed, {fail} could not be designed, "
                             f"over {len(pm)} markers design was attempted for"),
            ("Primer in-silico PCR", f"{st.count('pass')} one product; "
                                     f"{st.count('unverified')} NOT VERIFIED; "
                                     f"{st.count('danger')} DANGER. Only one product is a "
                                     f"pass: every other verdict, the default included, "
                                     f"means this pair has not been shown to amplify one "
                                     f"locus. No pair here is bench-validated (R3)."),
            ("Primer in-silico PCR caveat", ispcr.CAVEAT)]


def _nl_caveat(result):
    """The one sentence saying a model chose this variant, or None if the user named it.

    None means the user typed the identifier, and every format must then render nothing at
    all rather than a "none": an absent caveat is the honest rendering of an absent model.
    """
    prov = result.provenance
    model = prov.get("nl_model")
    if not model:
        return None
    return (f"Variant chosen by a language model ({model}) from the text: "
            f"{(prov.get('nl_text') or '').strip()!r}. The model was not given and did not "
            f"supply the coordinate, which was looked up live, but WHICH variant this panel "
            f"is about was its choice. Confirm it is the intended variant.")


def _het(maf):
    """2pq from a minor allele frequency. The one place this formula is written."""
    return None if maf is None else round(2 * maf * (1 - maf), 4)


def _rank_het(m, anc):
    """The ancestry-matched 2pq prior for this marker, or None where that population has
    no gnomAD frequency here (the engine then falls back to the global 2pq)."""
    return _het(m.per_pop_maf.get(anc)) if anc else None


def _decider(m, anc):
    """The number that decided this marker's rank.

    Mirrors panelbuilder._rank_key across a seam: the engine exposes no per-marker key.
    The self-check pins the two together against the engine's real key.
    """
    h = _rank_het(m, anc)
    return m.het if h is None else h


def _map_approx(result) -> bool:
    """True if any cM value came from the 1 cM/Mb fallback rather than the map."""
    return any(m.map_approx for m in result.candidates)


def _cm_note(result) -> str:
    src = result.provenance["sources"]["genetic_map"]
    if _map_approx(result):
        return (f"cM/recomb_fraction are APPROXIMATE: some markers fall outside the "
                f"genetic map and use a 1 cM/Mb fallback (see map_approx column). "
                f"Map: {src}")
    return f"cM/recomb_fraction from genetic map: {src}"


def _genomic(v) -> str:
    return f"chr{v.chrom}:{v.pos_grch38} {v.vcf_ref}>{v.vcf_alt} ({v.build})"


def _facts(result) -> list[tuple[str, str]]:
    """Ordered (label, value) pairs describing identity, evidence and provenance."""
    v, rar, prov, cov = result.variant, result.rarity, result.provenance, result.coverage
    src = prov["sources"]
    strand = {1: "plus (+1)", -1: "minus (-1)"}.get(v.strand, "unknown")
    f = [
        ("Reference build", v.build),
        ("Query", v.query),
        ("rsID", v.rsid or "n/a"),
        ("Gene", v.gene or "n/a"),
        # R7: the two forms are always shown side by side and labelled.
        ("Genomic (VCF)", _genomic(v)),
        ("Transcript sense (HGVS c.)", v.transcript_sense_change()),
        ("Gene strand", strand),
        ("ClinVar significance", v.clinical_significance or "n/a"),
        ("ClinVar review status", v.review_status or "n/a"),
        ("ClinVar accession", v.clinvar_accession or "n/a"),
    ]
    # Directly under Query, since it qualifies the query itself: the identity of the
    # variant, not a footnote about how it was typed.
    if caveat := _nl_caveat(result):
        f.insert(2, ("WARNING: model-chosen variant", caveat))
    if v.pos_grch37:
        # R6: carried for display only; every computation above is on pos_grch38.
        f.append(("GRCh37 position (display only)", f"chr{v.chrom}:{v.pos_grch37}"))
    if v.build_note:
        f.append(("Build note", v.build_note))
    f += [
        ("gnomAD genome AF", _num(rar.gnomad_af_genome)),
        ("gnomAD genome AC / AN", f"{_num(rar.gnomad_ac_genome)} / {_num(rar.gnomad_an_genome)}"),
        ("1000 Genomes AC", _num(rar.thousand_genomes_ac)),
        ("Population LD usable", str(rar.population_LD_usable)),
        ("LD verdict reason", rar.reason),
        ("Ranking key", _ranking_key(result)),
        ("Ranking exclusion (R2)",
         "LD with the pathogenic variant is never a ranking key."),
        ("Requested build", prov["requested_build"]),
        ("Window (bp, each side)", str(prov["window_bp"])),
        ("Common MAF floor", str(prov["common_maf"])),
        # Which ancestry was asked for, and nothing more: never name a 2pq here, the sort
        # key is _ranking_key's to report.
        ("Ancestry re-rank", prov["ancestry_rank"] or "none"),
        ("Candidates (common pool)", str(prov["candidate_n"])),
        ("Recommended panel size", str(len(result.recommended))),
        ("Coverage lower / higher coord", f"{cov['lower_count']} / {cov['higher_count']} markers; "
                               f"core-near {cov['lower_core_near']} / {cov['higher_core_near']}"),
        ("Coverage flags (R5)", "; ".join(cov["flags"]) or "none"),
        # Beside the coverage flags: both are per-side statements about the same shortlist.
        *_star_facts(result),
        *_primer_facts(result),
        ("Source: ClinVar", src["clinvar"]),
        ("Source: Ensembl", src["ensembl"]),
        ("Source: gnomAD", src["gnomad"]),
        ("Source: genetic map", src["genetic_map"]),
        ("Genetic map note", _cm_note(result)),
        ("Source data as of (UTC)", prov["queried_utc"]),
        ("Panel built (UTC)", prov.get("built_utc") or prov["queried_utc"]),
        ("Source responses", f"{prov.get('source_responses_from_network', 0)} fetched live, "
                             f"{prov.get('source_responses_from_cache', 0)} from local cache"),
        ("Build elapsed (s)", str(prov["elapsed_s"])),
    ]
    return f


def _num(x) -> str:
    return "" if x is None else repr(x) if isinstance(x, float) else str(x)


def _row(m, recommended_ids: set, anc=None, primers=False) -> list:
    """Values in _columns(anc, primers) order. Keep the two in step."""
    return ([m.rsid, m.chrom, m.pos, m.ref, m.alt, m.dist, m.side, m.tier,
             m.maf, m.af, m.an, m.het, m.het_max_pop]
            + ([_rank_het(m, anc)] if anc else [])
            + [m.per_pop_maf.get(p) for p in POPS]
            + [m.per_pop_an.get(p) for p in POPS]
            + [m.cm, m.recomb_fraction, m.hotspot_between, m.map_approx,
               m.ensembl_pos_check, m.variant_id in recommended_ids, _star_cell(m)]
            + (_primer_row(m) if primers else []))


def FILENAME(result, ext: str) -> str:
    v = result.variant
    stem = v.rsid or re.sub(r"\W+", "_", v.query).strip("_")[:40] or "variant"
    date = result.provenance["queried_utc"][:10]
    return f"originmarker_{stem}_{v.build}_{date}.{ext}"


# --------------------------------------------------------------------------- #
# CSV
# --------------------------------------------------------------------------- #

def to_csv(result) -> bytes:
    buf = io.StringIO(newline="")
    w = csv.writer(buf, lineterminator="\n")
    c = lambda s="": buf.write(f"# {s}\n".rstrip(" ") if s else "#\n")

    c("OriginMarker candidate linkage-marker panel - RESEARCH USE ONLY")
    c(pb.DISCLAIMER)                                   # R8, verbatim
    c()
    for k, val in _facts(result):
        c(f"{k}: {val}")
    c()
    c("CANDIDATE MARKERS ONLY (R3). These are not usable as-is:")
    for i, step in enumerate(pb.LAYER_B_STEPS, 1):     # R3, verbatim
        c(f"  {i}. {step}")
    c()
    c("het_2pq_prior_* are POPULATION PRIORS (R4): 2pq is the expected heterozygosity")
    c("in the population, NOT evidence that this carrier is heterozygous.")
    if _rank_pop(result):
        anc_ = _rank_pop(result)
        c(f"het_2pq_prior_{anc_} is the number the ranking keyed on. It is empty where gnomAD")
        c(f"reports no {anc_} frequency at that site (see an_{anc_}), and the ranking then")
        c("fell back to het_2pq_prior_global for that marker.")
    c("signed_dist_bp is relative to the pathogenic variant; negative = lower GRCh38 coordinate.")
    c("side is the genomic axis (lower / higher GRCh38 coordinate). It is NOT an arm label:")
    c("nothing here knows where the centromere is, so side never names a chromosome arm.")
    c(f"gnomad_an / an_* are allele NUMBERS: what each frequency rests on. Sites below")
    c(f"AN {pb.CALL_RATE_AN_FLOOR} are excluded entirely; a population MAF is reported only")
    c(f"above AN {pb.MIN_POP_AN} (~{pb.MIN_POP_AN // 2} people). gnomAD QC-failed sites")
    c("(AC0 / AS_VQSR) are excluded and never appear here.")
    if _has_primers(result):
        c("primer_* are CANDIDATE designs: no pair here has been tested at the bench.")
        c("primer_fwd/primer_rev read 5' to 3' as they would be ordered.")
        c(f"primer_insilico_pcr is the verdict. ONLY {ispcr.ONE_PRODUCT!r} is a pass. The")
        c(f"default {ispcr.NOT_CHECKED!r} means nothing looked for other products in the")
        c(f"genome, {ispcr.DANGER!r} means in-silico PCR found something wrong with the pair,")
        c(f"and {ispcr.UNKNOWN!r} means the answer could not be read. Read primer_warnings")
        c("before ordering any pair: it is never empty for a pair that exists.")
        c("primer_design_error carries the reason where no pair could be designed. Empty")
        c("primer_* cells mean no design was attempted for that marker, which is not a")
        c("verdict about it.")
        c(ispcr.CAVEAT)
    c(f"All positions are {result.variant.build} (R6).")
    c(_cm_note(result))
    c()

    anc = _rank_pop(result)
    # A column, not just the header block above: pandas.read_csv(comment='#') strips every
    # line of that block, and the reader who does that is exactly the one who must not lose
    # this. The model id rides every row because the column, not a footnote, is what
    # survives; the full caveat stays in the header block for a human.
    nl_model = result.provenance.get("nl_model")
    extra = ["variant_chosen_by_language_model"] if nl_model else []
    rec = {m.variant_id for m in result.recommended}
    pr = _has_primers(result)
    w.writerow(_columns(anc, pr) + extra)
    for m in result.candidates:
        w.writerow(_row(m, rec, anc, pr) + ([nl_model] if nl_model else []))
    return buf.getvalue().encode("utf-8")


# --------------------------------------------------------------------------- #
# JSON
# --------------------------------------------------------------------------- #

def to_json(result) -> bytes:
    d = result.to_dict()
    rec = {m.variant_id for m in result.recommended}
    anc = _rank_pop(result)
    pr = _has_primers(result)
    # Both lists, because to_dict() writes the shortlist out twice: once inside `candidates`
    # and again whole under `recommended`. Enriching only the first left the same pair in two
    # different shapes in one file, and a reader who took the shortlist rather than filtering
    # the candidates got the lesser one.
    # zip relies on to_dict() being asdict(), which preserves list order; the assert checks it.
    for key in ("candidates", "recommended"):
        for m, md in zip(getattr(result, key), d[key]):
            assert md["variant_id"] == m.variant_id, f"to_dict() reordered {key}"
            md["in_recommended_panel"] = md["variant_id"] in rec
            # STAR_FIELD needs no line here: it is a Marker field, so asdict() already carried
            # it, and provenance["flanking_criteria"] already carries the engine's words for it.
            # Via _rank_het, the same expression the other three formats use: only _rank_het is
            # pinned to the engine's ranking key by the self-check.
            if anc:
                md[f"het_2pq_prior_{anc}"] = _rank_het(m, anc)
            # Always, because Marker.primer is a field and asdict() has already put a null
            # here: the key is the engine's shape and a machine reader keeps it. Null means no
            # design for this marker, which is not a verdict about it. The PROSE below is what
            # stays conditional, since explaining primers to a panel that has none explains
            # nothing.
            md["primer"] = _primer_dict(m)
    prov = dict(d["provenance"], genetic_map_note=_cm_note(result),
                ranking_key=_ranking_key(result),
                ranking_excludes="LD with the pathogenic variant is never a ranking key (R2)",
                het_2pq_semantics="population prior, not a per-carrier genotype claim (R4)",
                transcript_sense=result.variant.transcript_sense_change())
    if _has_primers(result):
        prov["primer_semantics"] = (
            f"Candidate designs, never bench-validated (R3). Of the four "
            f"primer_insilico_pcr verdicts only {ispcr.ONE_PRODUCT!r} is a pass, and it is "
            f"an in-silico check against the reference rather than a wet-lab result. "
            f"{ispcr.NOT_CHECKED!r} is the default, {ispcr.DANGER!r} is a finding to act on "
            f"and {ispcr.UNKNOWN!r} is an unreadable answer; all three mean the pair has not "
            f"been shown to amplify one locus. primer_warnings is never empty for a pair.")
        prov["primer_insilico_pcr_caveat"] = ispcr.CAVEAT
    # nl_text/nl_model are already here from build(), stable shape, null when the user
    # named the variant. The rendered sentence is added only when there IS one to render.
    if caveat := _nl_caveat(result):
        prov["nl_caveat"] = caveat
    payload = {
        "provenance": prov,
        "variant": d["variant"],
        "rarity": d["rarity"],
        "coverage": d["coverage"],
        "candidates": d["candidates"],
        "recommended": d["recommended"],
        "layer_b_steps": list(pb.LAYER_B_STEPS),
        "disclaimer": pb.DISCLAIMER,
    }
    return json.dumps(payload, indent=2, ensure_ascii=False).encode("utf-8")


# --------------------------------------------------------------------------- #
# The mark. web/public/favicon.svg is the ONE definition of its geometry: the PDF
# masthead and the XLSX render that file, so neither can drift from the site's
# monogram. Nothing here may restate its path, its transform or its colours.
# --------------------------------------------------------------------------- #

# public/ in the repo; dist/ in the container image, where vite has copied public/ in.
MARK_SVG = [pathlib.Path(__file__).resolve().parent.parent / "web" / d / "favicon.svg"
            for d in ("public", "dist")]
_SVG_NS = "{http://www.w3.org/2000/svg}"


def _svg_polys(d: str) -> list[list[tuple[float, float]]]:
    """Sub-paths of an SVG path built from straight lines, absolute commands only.

    Raises on any command it cannot reproduce exactly: a curve flattened to a chord would
    be a silently wrong mark, and a wrong mark is worse than no mark.
    """
    toks = re.findall(r"[A-Za-z]|-?[\d.]+", d)
    polys, cur, x, y, op, i = [], [], 0.0, 0.0, None, 0
    while i < len(toks):
        if toks[i].isalpha():
            op, i = toks[i], i + 1
            if op in ("Z", "M") and cur:
                polys.append(cur)
                cur = []
            if op == "Z":
                continue
        if op not in ("M", "L", "H", "V"):
            raise ValueError(f"favicon.svg path: cannot draw command {op!r} exactly")
        if op in ("M", "L"):
            x, y, i, op = float(toks[i]), float(toks[i + 1]), i + 2, "L"   # M then implicit L
        elif op == "H":
            x, i = float(toks[i]), i + 1
        else:
            y, i = float(toks[i]), i + 1
        cur.append((x, y))
    if cur:
        polys.append(cur)
    return polys


def _mark() -> dict | None:
    """The monogram's geometry and colours, read from favicon.svg. None if it is absent.

    Not a general SVG reader: it takes the ring and the straight-line glyph outline that
    file holds, and raises on anything else rather than guess.
    """
    src = next((p for p in MARK_SVG if p.exists()), None)
    if src is None:
        return None
    svg = ElementTree.fromstring(src.read_bytes())
    vx, vy, vw, vh = (float(t) for t in svg.get("viewBox").split())
    assert vw == vh, f"favicon.svg viewBox {vw}x{vh} is not square: the ring is a circle"
    circle, path = svg.find(_SVG_NS + "circle"), svg.find(_SVG_NS + "path")
    t = re.fullmatch(r"translate\(([-\d.]+),([-\d.]+)\)\s*scale\(([-\d.]+),([-\d.]+)\)",
                     (path.get("transform") or "translate(0,0) scale(1,1)").strip())
    assert t, f"favicon.svg: unreadable path transform {path.get('transform')!r}"
    return {"viewbox": (vx, vy, vw, vh),
            "circle": tuple(float(circle.get(k)) for k in ("cx", "cy", "r", "stroke-width")),
            "ring": circle.get("stroke"),
            "fill": path.get("fill"),
            "transform": tuple(float(g) for g in t.groups()),
            "polys": _svg_polys(path.get("d"))}


def _mark_png(px: int) -> bytes | None:
    """The monogram as a transparent PNG of px square, for formats that take no vector.

    Pillow draws no antialiased edge, hence the supersample. It arrives with reportlab
    (a hard dependency of it), so this adds nothing to requirements.txt.
    """
    from PIL import Image, ImageDraw
    m = _mark()
    if m is None:
        return None
    SS = 4
    vx, vy, vw, _ = m["viewbox"]
    k = px * SS / vw
    at = lambda ux, uy: ((ux - vx) * k, (uy - vy) * k)     # viewBox units -> pixels, y down
    img = Image.new("RGBA", (px * SS, px * SS), (0, 0, 0, 0))
    d = ImageDraw.Draw(img)
    cx, cy, r, sw = m["circle"]
    # PIL strokes inward from the bounding box; SVG centres the stroke on r.
    d.ellipse([*at(cx - r - sw / 2, cy - r - sw / 2), *at(cx + r + sw / 2, cy + r + sw / 2)],
              outline=m["ring"], width=round(sw * k))
    tx, ty, sx, sy = m["transform"]
    for poly in m["polys"]:
        d.polygon([at(tx + sx * gx, ty + sy * gy) for gx, gy in poly], fill=m["fill"])
    out = io.BytesIO()
    img.resize((px, px), Image.LANCZOS).save(out, "PNG")
    return out.getvalue()


# --------------------------------------------------------------------------- #
# XLSX
# --------------------------------------------------------------------------- #

def to_xlsx(result) -> bytes:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image
    from openpyxl.styles import Alignment, Font

    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    rec = {m.variant_id for m in result.recommended}
    anc = _rank_pop(result)
    pr = _has_primers(result)

    wb = Workbook()
    wb.remove(wb.active)

    def sheet(title):
        """New sheet with pb.DISCLAIMER pinned in row 1 (R8, on EVERY sheet)."""
        ws = wb.create_sheet(title)
        ws["A1"] = pb.DISCLAIMER
        ws["A1"].font = Font(bold=True, italic=True)
        return ws

    def table(title, markers):
        ws = sheet(title)
        headers = _columns(anc, pr)
        if _map_approx(result):
            headers[headers.index("cm_to_variant")] = "cm_to_variant (APPROX - see map_approx)"
        ws.append([])                       # row 2 spacer keeps A1 readable
        ws.append(headers)                  # row 3
        for c in ws[3]:
            c.font = bold
        for m in markers:
            ws.append(_row(m, rec, anc, pr))
        ws.freeze_panes = "A4"              # rows 1-3: disclaimer + header stay visible
        _autosize(ws)
        return ws

    ws = table("Recommended panel", result.recommended)
    # First sheet only, for the same reason the PDF marks only page 1. An openpyxl image
    # floats over the grid rather than filling a cell, so it goes in a gutter opened by
    # indenting A1's DISPLAY: the disclaimer's text is untouched (R8) and stays uncovered.
    if png := _mark_png(54):
        img = Image(io.BytesIO(png))
        img.width = img.height = 18
        img.anchor = "A1"
        ws.add_image(img)
        # horizontal is explicit because Excel drops an indent under General alignment.
        ws["A1"].alignment = Alignment(horizontal="left", indent=3)
    ws["A2"] = (f"Balanced subset, both sides (R5). CANDIDATES ONLY - genotype the carrier, "
                f"keep heterozygous markers, then phase (R3). het_2pq_prior_* are population "
                f"priors (R4). Positions {result.variant.build} (R6)."
                + (" primer_* are candidate designs, never bench-validated: "
                   "primer_insilico_pcr='not_checked' is the DEFAULT and is NOT a pass."
                   if pr else ""))
    table("All candidates", result.candidates)["A2"] = (
        f"Full common pool (MAF >= {result.provenance['common_maf']}), ranked on "
        f"{_ranking_key(result)} - never on LD with the pathogenic variant (R2). "
        f"side is the genomic axis, not a chromosome arm: nothing here knows where the "
        f"centromere is. Positions {result.variant.build} (R6).")

    ws = sheet("Variant + provenance")
    ws.append([])
    ws.append(["Field", "Value"])
    for c in ws[3]:
        c.font = bold
    for k, v in _facts(result):
        ws.append([k, v])
    for row in ws.iter_rows(min_row=4, min_col=2, max_col=2):
        row[0].alignment = wrap
    ws.freeze_panes = "A4"
    _autosize(ws, cap=95)

    # Sheet name avoids "Layer B": that is this repo's word for the wet-lab half, not the
    # field's, and a reader opening the workbook has never seen the spec.
    ws = sheet("Using these markers")
    ws.append([])
    ws.append(["Step", "Action"])
    for c in ws[3]:
        c.font = bold
    for i, step in enumerate(pb.LAYER_B_STEPS, 1):
        ws.append([i, step])
    ws.append([])
    ws.append(["", "This app builds CANDIDATE panels only. It cannot determine phase, "
                   "and it makes no claim that any given carrier is heterozygous at any "
                   "marker - 2pq is a population prior (R3/R4)."])
    for row in ws.iter_rows(min_row=4, min_col=2, max_col=2):
        row[0].alignment = wrap
    ws.freeze_panes = "A4"
    _autosize(ws, cap=110)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _autosize(ws, cap=28):
    """Width from the longest cell, ignoring the pinned disclaimer/note rows."""
    for col in ws.columns:
        letter = col[0].column_letter
        widest = max((len(str(c.value)) for c in col[2:] if c.value is not None),
                     default=10)
        ws.column_dimensions[letter].width = min(max(widest + 2, 9), cap)


# --------------------------------------------------------------------------- #
# PDF
# --------------------------------------------------------------------------- #

# U+2605, the engine's own glyph, and ZapfDingbats is the only standard PDF font that has
# it. Hand reportlab the CHARACTER and it encodes it to that font's code point; hand it
# that code point ("H") and it silently draws a different dingbat.
STAR = "★"


def _star_glyph(c, x, y, size) -> None:
    """The star, baseline-left at (x, y). Shape only: it takes the caller's fill colour.

    This page gets printed and filed, so the star has to carry its meaning in its outline.
    A hue is the first thing a black-and-white printer throws away.
    """
    c.setFont("ZapfDingbats", size)
    c.drawString(x, y, STAR)


def to_pdf(result) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfbase.pdfmetrics import stringWidth
    from reportlab.pdfgen import canvas as rl_canvas

    W, H = letter
    L, R, TOP, BOT = 40, W - 40, H - 42, 78      # margins; BOT clears the footer
    v = result.variant
    out = io.BytesIO()
    c = rl_canvas.Canvas(out, pagesize=letter)
    c.setTitle(FILENAME(result, "pdf"))
    state = {"y": TOP, "page": 0}

    def footer():
        """R8: pb.DISCLAIMER verbatim in the footer of EVERY page."""
        c.setFont("Helvetica-Oblique", 7)
        c.setFillColor(colors.black)
        yy = 44
        for line in _wrap(pb.DISCLAIMER, "Helvetica-Oblique", 7, R - L, stringWidth):
            c.drawString(L, yy, line)
            yy -= 8.5
        c.setFont("Helvetica", 6.5)
        c.setFillColor(colors.grey)
        c.drawString(L, 26, f"OriginMarker | {v.build} | queried {result.provenance['queried_utc']} "
                            f"| {result.provenance['sources']['gnomad']} | page {state['page']}")
        c.setFillColor(colors.black)

    def new_page():
        if state["page"]:
            footer()
            c.showPage()
        state["page"] += 1
        state["y"] = TOP

    def need(h):
        if state["y"] - h < BOT:
            new_page()

    def text(s, size=8, font="Helvetica", gap=2, colour=colors.black, x0=L):
        need(size + gap)
        c.setFont(font, size)
        c.setFillColor(colour)
        for line in _wrap(s, font, size, R - x0, stringWidth):
            need(size + gap)
            c.drawString(x0, state["y"] - size, line)
            state["y"] -= size + gap
        c.setFillColor(colors.black)

    new_page()

    # --- masthead -----------------------------------------------------------
    # Page 1 only, and never inside new_page(): a mark at the top of the report is a
    # masthead, a mark on every page is noise the reader has to look past.
    MARK = 26
    if _draw_mark(c, L, state["y"], MARK):
        c.setFont("Helvetica-Bold", 14)
        c.drawString(L + MARK + 8, state["y"] - 18,
                     "OriginMarker - candidate linkage markers for PGT-M")
        state["y"] -= MARK + 4
    else:
        text("OriginMarker - candidate linkage markers for PGT-M", 14, "Helvetica-Bold", 4)

    # --- variant card -------------------------------------------------------
    text(f"{v.rsid or v.query}   {v.gene or ''}   {v.clinical_significance or ''}"
         f"   {v.clinvar_accession or ''}", 9, "Helvetica-Bold", 5)
    state["y"] -= 2
    card = [
        # R7: both forms, always labelled.
        ("Genomic (VCF):", _genomic(v)),
        ("Transcript sense (HGVS c.):", v.transcript_sense_change()),
        ("Query:", v.query),
        ("ClinVar review:", v.review_status or "n/a"),
        # fmt_af, not _num: this is prose, and the exact value sits beside it as AC/AN.
        # Machine-readable fields (the CSV data columns) must never be formatted this way.
        ("gnomAD genome AF:", f"{pb.fmt_af(result.rarity.gnomad_af_genome)}  "
                              f"(AC {_num(result.rarity.gnomad_ac_genome)} / "
                              f"AN {_num(result.rarity.gnomad_an_genome)})"),
        ("1000G AC:", _num(result.rarity.thousand_genomes_ac)),
        ("Population LD usable:", f"{result.rarity.population_LD_usable} - {result.rarity.reason}"),
        ("Window / MAF floor:", f"+/-{result.provenance['window_bp']} bp | MAF >= "
                                f"{result.provenance['common_maf']} | ancestry "
                                f"{_rank_pop(result) or 'none'}"),
        # tel_/cen_ are coverage-dict key names only: the label a reader sees names the
        # genomic axis, never a chromosome arm.
        ("Coverage (R5):", f"lower coord {result.coverage['lower_count']} / higher coord "
                           f"{result.coverage['higher_count']}"
                           f" | flags: {'; '.join(result.coverage['flags']) or 'none'}"),
    ]
    kw = 118
    for k, val in card:
        need(11)
        c.setFont("Helvetica-Bold", 8)
        c.drawString(L, state["y"] - 8, k)
        c.setFont("Helvetica", 8)
        for i, line in enumerate(_wrap(str(val), "Helvetica", 8, R - L - kw, stringWidth)):
            if i:
                need(10)
            c.drawString(L + kw, state["y"] - 8, line)
            state["y"] -= 10
    # On the card at body size and in red, never a footnote: this is the same class of fact
    # as the disclaimer, and this is the page that gets printed and filed.
    if caveat := _nl_caveat(result):
        state["y"] -= 4
        text(caveat, 8, "Helvetica-Bold", 2.5, colors.HexColor("#b12222"))
    text(f"Ranked on {_ranking_key(result)}. LD with the pathogenic variant is never a "
         f"ranking key (R2).", 7, "Helvetica-Oblique", 3, colors.grey)

    # --- locus figure -------------------------------------------------------
    state["y"] -= 6
    _figure(c, result, L, R, state, need, colors)

    # --- recommended table --------------------------------------------------
    state["y"] -= 10
    text(f"Recommended panel ({len(result.recommended)} candidate markers, both sides)",
         10, "Helvetica-Bold", 4)
    approx = " *approx" if _map_approx(result) else ""
    anc = _rank_pop(result)
    # The head must name the population the deciding 2pq is for; a bare "2pq prior" on a
    # page stamped with an ancestry reads as the global figure. The global stays alongside,
    # since it is the fallback.
    dec_col = f"2pq {anc}" if anc else "2pq global"
    # The star column exists only when the page carries a star to explain. It leads, so the
    # eye finds it down the left edge without reading eleven columns first.
    star_legend = _star_legend(result, result.recommended)
    cols = (([(STAR, 14, "l")] if star_legend else [])
            + [("rsID", 58, "l"), ("pos " + v.build, 58, "r"), ("dist bp", 46, "r"),
               ("side", 42, "l"), ("tier", 52, "l"), ("2pq global", 42, "r")]
            + ([(dec_col, 46, "r")] if anc else [])
            + [("MAF", 32, "r"), ("cM" + approx, 36, "r"),
               ("hotspot", 34, "l"), ("pos chk", 72, "l")])
    assert sum(w for _, w, _ in cols) <= R - L, "PDF table wider than the page"

    def header_row():
        need(14)
        c.setFont("Helvetica-Bold", 6.8)
        x = L
        for name, w_, al in cols:
            if name == STAR:
                _star_glyph(c, x, state["y"] - 7, 6.8)
                c.setFont("Helvetica-Bold", 6.8)
            else:
                c.drawString(x, state["y"] - 7, name)
            x += w_
        state["y"] -= 9
        c.setLineWidth(0.4)
        c.line(L, state["y"], R, state["y"])
        state["y"] -= 2

    header_row()
    for m in sorted(result.recommended, key=lambda m: m.dist):
        if state["y"] - 9 < BOT:
            new_page()
            text(f"Recommended panel (cont.) - {v.build}", 10, "Helvetica-Bold", 4)
            header_row()
        # Render m.side; never derive an arm from the sign of m.dist. Nothing here knows
        # where the centromere is, and below a q-arm locus is centromeric, not telomeric.
        rh = _rank_het(m, anc)
        vals = (([STAR if _starred(m) else ""] if star_legend else [])
                + [m.rsid, f"{m.pos:,}", f"{m.dist:+,}", m.side, m.tier, f"{m.het:.3f}"]
                + ([("-" if rh is None else f"{rh:.3f}")] if anc else [])
                + [f"{m.maf:.3f}",
                   "n/a" if m.cm is None else f"{m.cm:.4f}",
                   "yes" if m.hotspot_between else "no", m.ensembl_pos_check or "-"])
        c.setFont("Helvetica", 6.8)
        c.setFillColor(colors.red if m.ensembl_pos_check and
                       m.ensembl_pos_check.startswith("MISMATCH") else colors.black)
        x = L
        for (name, w_, al), val in zip(cols, vals):
            if name == STAR:
                # The row's own colour, so a disputed row's star is as red as its warning.
                if val:
                    _star_glyph(c, x, state["y"] - 7, 7)
                    c.setFont("Helvetica", 6.8)
            elif al == "r":
                c.drawRightString(x + w_ - 6, state["y"] - 7, str(val))
            else:
                c.drawString(x, state["y"] - 7, str(val))
            x += w_
        c.setFillColor(colors.black)
        state["y"] -= 9
    note = (f"2pq = expected heterozygosity in the population (R4), NOT evidence that this "
            f"carrier is heterozygous. {dec_col} is the figure the ranking keyed on. "
            f"dist is signed from the variant; - = lower coordinate. side is the genomic "
            f"axis, not a chromosome arm: nothing here knows where the centromere is.")
    if anc:
        note += (f" A dash means gnomAD reports no {anc} frequency at that site, and the "
                 f"ranking fell back to the global 2pq for that marker.")
    if _map_approx(result):
        note += " *cM approximate: 1 cM/Mb fallback outside the map."
    text(note, 6.5, "Helvetica-Oblique", 2, colors.grey)

    # --- star legend --------------------------------------------------------
    # Black, not grey: this defines a symbol printed beside a marker, so it outranks the
    # note above it. The engine's paragraphs open with a literal U+2605 that no standard PDF
    # text font carries, so the glyph is drawn and the words are rendered from after it.
    if star_legend:
        state["y"] -= 6
        for i, para in enumerate(star_legend):
            need(24)                       # keeps the glyph with its own first line
            if not i:
                _star_glyph(c, L, state["y"] - 7, 7.5)
            text(para.lstrip("★ "), 6.4, "Helvetica-Oblique", 1.7, colors.black, L + 11)
        text(f"Markers meeting the flanking criteria: {_star_count_text(result)}",
             6.4, "Helvetica-Oblique", 1.7, colors.black, L + 11)

    # --- primers ------------------------------------------------------------
    # Before the lab steps and after the table: these are what a reader hands to a bench,
    # and the steps below are what the bench then has to do with them.
    _primer_section(c, result, L, R, TOP - BOT, state, need, text, colors, stringWidth)

    # --- the wet-lab hand-off -----------------------------------------------
    state["y"] -= 8
    need(30)
    text("Using these markers: lab steps this app cannot do for you",
         10, "Helvetica-Bold", 4)
    for i, step in enumerate(pb.LAYER_B_STEPS, 1):     # R3, verbatim
        need(11)
        c.setFont("Helvetica-Bold", 7.6)
        c.drawString(L, state["y"] - 8, f"{i}.")
        c.setFont("Helvetica", 7.6)
        for j, line in enumerate(_wrap(step, "Helvetica", 7.6, R - L - 14, stringWidth)):
            if j:
                need(9.6)
            c.drawString(L + 14, state["y"] - 8, line)
            state["y"] -= 9.6
    text(f"Sources: ClinVar {result.provenance['sources']['clinvar']} | Ensembl "
         f"{result.provenance['sources']['ensembl']} | gnomAD "
         f"{result.provenance['sources']['gnomad']} | map "
         f"{result.provenance['sources']['genetic_map']} | queried "
         f"{result.provenance['queried_utc']} | build {v.build} (R6)",
         6.3, "Helvetica", 2, colors.grey)

    footer()
    c.showPage()
    c.save()
    return out.getvalue()


def _draw_mark(c, x, y_top, size) -> bool:
    """Draw the monogram from favicon.svg, top-left at (x, y_top). False if it is absent.

    Points are baked to the page rather than pushed through a CTM, so the coordinates the
    self-check reads back out of the PDF are the ones the reader sees.
    """
    from reportlab.lib import colors
    m = _mark()
    if m is None:
        return False
    vx, vy, vw, _ = m["viewbox"]
    k = size / vw
    at = lambda ux, uy: (x + (ux - vx) * k, y_top - (uy - vy) * k)   # SVG y is down, PDF's up
    cx, cy, r, sw = m["circle"]
    c.setStrokeColor(colors.HexColor(m["ring"]))
    c.setLineWidth(sw * k)
    c.circle(*at(cx, cy), r * k, stroke=1, fill=0)
    tx, ty, sx, sy = m["transform"]
    p = c.beginPath()
    for poly in m["polys"]:
        for i, (gx, gy) in enumerate(poly):
            (p.moveTo if i == 0 else p.lineTo)(*at(tx + sx * gx, ty + sy * gy))
        p.close()
    c.setFillColor(colors.HexColor(m["fill"]))
    c.drawPath(p, stroke=0, fill=1)
    c.setFillColor(colors.black)
    return True


# Sequences are set in Courier so the two 5' ends line up under one another and a reader can
# count bases off the printed page.
SEQ_FONT, SEQ_SIZE = "Courier-Bold", 7.6


def _para(c, s, font, size, x, R, y, colour, stringWidth) -> float:
    """One wrapped paragraph from y downwards; returns the y it ends at. c=None measures."""
    for line in _wrap(s, font, size, R - x, stringWidth):
        if c:
            c.setFont(font, size)
            c.setFillColor(colour)
            c.drawString(x, y - size, line)
        y -= size + 1.8
    return y


def _verdict_box(c, head, paras, kind, x, R, y, colors, stringWidth) -> float:
    """The verdict and every warning under it, from y down; returns the y it ends at.

    c=None measures. This page is printed in black and white, so the three states differ in
    ink, outline and weight, never in hue alone: a danger is reversed out of a filled bar,
    an unverified pair is boxed and bold, and only a verified pair is left plain.
    """
    size, pad = 6.8, 4
    font = "Helvetica" if kind == "plain" else "Helvetica-Bold"
    lines = []
    for i, para in enumerate(paras):
        # A colon, because the engine's own warnings already open at a shout: without one,
        # "NOT VERIFIED NOT CHECKED AGAINST THE GENOME" reads as a stutter rather than as a
        # label over the engine's sentence.
        lines += _wrap(para if i else f"{head}: {para}", font, size,
                       R - x - 2 * pad - 4, stringWidth)
    h = len(lines) * (size + 2) + 2 * pad
    if c:
        if kind == "danger":
            c.setFillColor(colors.black)
            c.rect(x, y - h, R - x, h, stroke=0, fill=1)
            c.setFillColor(colors.white)
        else:
            if kind == "box":
                c.setStrokeColor(colors.black)
                c.setLineWidth(0.9)
                c.rect(x, y - h, R - x, h, stroke=1, fill=0)
            c.setFillColor(colors.black)
        c.setFont(font, size)
        yy = y - pad
        for line in lines:
            c.drawString(x + pad + 2, yy - size, line)
            yy -= size + 2
        c.setFillColor(colors.black)
    return y - h


def _product_text(p, m) -> str:
    if p.product_start is None or p.product_end is None:
        return "product coordinates: not reported by this build"
    return (f"product {p.product_size:,} bp   chr{m.chrom}:{p.product_start:,}-"
            f"{p.product_end:,}")


def _primer_block(c, m, L, R, y, colors, stringWidth) -> float:
    """One marker's primer pair, from y downwards; returns the y it ends at.

    c=None measures without drawing. The space a block reserves and the ink it lays down come
    out of this one pass, so a block cannot be reserved shorter than it draws and split
    across a page in the middle of its own warning.
    """
    p, x = _primer(m), L + 9
    if c:
        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 8.2)
        c.drawString(L, y - 8, m.rsid)
        c.setFont("Helvetica", 7.2)
        c.drawString(L + 62, y - 8, f"chr{m.chrom}:{m.pos:,} {m.ref}>{m.alt}")
        c.drawRightString(R, y - 8, _product_text(p, m))
    y -= 11.5

    # A pair with no oligos is a failure whether or not it carries the reason, and it says so
    # rather than printing an empty block that reads as a pair with nothing wrong.
    if p.error or p.fwd is None or p.rev is None:
        return _verdict_box(c, "NO PRIMER DESIGNED",
                            [str(p.error or "This build reports no pair and no reason.")],
                            "box", x, R, y, colors, stringWidth)

    for tag, o in (("FWD", p.fwd), ("REV", p.rev)):
        s = f"{tag} 5'-{o.seq}-3'"
        met = f"Tm {o.tm:.1f} C   GC {o.gc:.1f}%   {o.length} nt"
        # A long oligo takes the metrics to their own line rather than printing them over it.
        drop = 0 if (stringWidth(s, SEQ_FONT, SEQ_SIZE)
                     + stringWidth(met, "Helvetica", 6.8) + 16 <= R - x) else 9
        if c:
            c.setFillColor(colors.black)
            c.setFont(SEQ_FONT, SEQ_SIZE)
            c.drawString(x, y - SEQ_SIZE, s)
            c.setFont("Helvetica", 6.8)
            c.drawRightString(R, y - SEQ_SIZE - drop, met)
        y -= SEQ_SIZE + 2.6 + drop
    # The engine's mask sentence stands on its own: a label in front of it would be this
    # file's word for a reach only the engine knows.
    y = _para(c, _mask_note(p), "Helvetica", 6.4, x, R, y - 1, colors.black, stringWidth)
    kind, head = {"danger": ("danger", "DANGER"),
                  "unverified": ("box", "NOT VERIFIED"),
                  "pass": ("plain", "VERIFIED CLEAN (in silico)")}[_pcr_state(p)]
    return _verdict_box(c, head, _primer_warnings(p), kind, x, R, y - 1.5,
                        colors, stringWidth)


def _primer_section(c, result, L, R, usable, state, need, text, colors, stringWidth):
    """A block per marker the engine designed a pair for, in coordinate order.

    Which markers get one is the engine's; this renders a block for every pair it finds and
    picks no marker of its own.
    """
    pm = [m for m in sorted(result.candidates, key=lambda m: m.dist) if _primer(m)]
    if not pm:
        return
    state["y"] -= 10
    # Keep the heading with its first block: a section title alone at the foot of a page
    # sends the reader looking for a section that is not there. The reserve overshoots the
    # heading and its lead on purpose, since reserving too much only moves the section down
    # a page and can never clip it.
    need(46 + state["y"] - _primer_block(None, pm[0], L, R, state["y"], colors, stringWidth))
    text(f"Primer pairs ({len(pm)} candidate designs, {result.variant.build})",
         10, "Helvetica-Bold", 4)
    text("Candidate designs: no pair here has been tested at the bench, and a pair was "
         "checked against the genome only where its verdict below says so. Both sequences "
         "read 5' to 3' as they would be ordered. Product coordinates are "
         f"{result.variant.build} (R6).", 6.5, "Helvetica-Oblique", 3, colors.grey)
    for m in pm:
        h = state["y"] - _primer_block(None, m, L, R, state["y"], colors, stringWidth)
        # Measured whole, reserved whole: a block never splits across a page. Nothing here
        # breaks one that is taller than a page, and no warning is ever shortened to fit.
        assert h <= usable, (f"the primer block for {m.rsid} is {h:.0f}pt on a {usable:.0f}pt "
                             f"page and would clip its own warning")
        need(h + 5)
        state["y"] = _primer_block(c, m, L, R, state["y"], colors, stringWidth) - 5


def _figure(c, result, L, R, state, need, colors):
    """Locus map: axis, dashed line at the variant, lollipops for recommended markers.

    Stem height comes from _decider, so the picture matches the table's deciding 2pq
    column. Colour = side.
    """
    # Bands, top-down: title, legend, stems, axis, ticks. axis_y is low enough that a
    # full-height stem (2pq=0.5) stops short of the legend row.
    FH = 104                                 # figure block height
    need(FH)
    top = state["y"]
    legend_y = top - 20
    axis_y = top - FH + 28
    stem_h = 42
    win = result.provenance["window_bp"]
    span = max([win] + [abs(m.dist) for m in result.recommended]) or 1
    mid = (L + R) / 2
    half = (R - L) / 2 - 14
    x_of = lambda d: mid + (d / span) * half

    anc = _rank_pop(result)
    dec_col = f"2pq {anc}" if anc else "2pq global"
    c.setFont("Helvetica-Bold", 8)
    c.setFillColor(colors.black)
    c.drawString(L, top - 8, f"Locus map ({result.variant.build}) - lollipop height = "
                             f"{dec_col} population prior (R4), colour = side")

    lo, hi = colors.HexColor("#2c6fbb"), colors.HexColor("#d1741f")
    # Keyed on m.side, not the sign of dist. An unknown side goes grey rather than
    # silently taking the other side's colour.
    side_col = {"lower coord": lo, "higher coord": hi}
    for m in sorted(result.recommended, key=lambda m: _decider(m, anc)):
        x = x_of(m.dist)
        h = (_decider(m, anc) / 0.5) * stem_h    # 2pq maxes out at 0.5
        col = side_col.get(m.side, colors.grey)
        c.setStrokeColor(col)
        c.setLineWidth(0.7)
        c.line(x, axis_y, x, axis_y + h)
        c.setFillColor(col)
        c.circle(x, axis_y + h, 2.1, stroke=0, fill=1)

    c.setStrokeColor(colors.black)
    c.setLineWidth(0.8)
    c.line(L, axis_y, R, axis_y)
    for d in (-span, -span / 2, 0, span / 2, span):
        x = x_of(d)
        c.setLineWidth(0.6)
        c.line(x, axis_y, x, axis_y - 3)
        c.setFont("Helvetica", 6)
        c.setFillColor(colors.grey)
        c.drawCentredString(x, axis_y - 11, f"{d/1000:+.0f} kb" if d else "0")

    c.setStrokeColor(colors.HexColor("#b12222"))
    c.setLineWidth(1.0)
    c.setDash(2, 2)
    c.line(x_of(0), axis_y - 5, x_of(0), axis_y + stem_h + 6)
    c.setDash()
    c.setFillColor(colors.HexColor("#b12222"))
    c.setFont("Helvetica-Bold", 6.5)
    c.drawCentredString(x_of(0), legend_y, result.variant.rsid or "pathogenic variant")

    c.setFont("Helvetica", 6)
    c.setFillColor(lo)
    c.drawString(L, legend_y, "lower GRCh38 coordinate")
    c.setFillColor(hi)
    c.drawRightString(R, legend_y, "higher GRCh38 coordinate")
    c.setFillColor(colors.grey)
    c.drawString(L, axis_y - 21, f"{dec_col} scale: 0 to 0.50 (max for a biallelic SNP)")
    c.setFillColor(colors.black)
    state["y"] = axis_y - 26


def _wrap(text: str, font: str, size: float, width: float, stringWidth) -> list[str]:
    lines, cur = [], ""
    for word in text.split():
        trial = f"{cur} {word}".strip()
        if stringWidth(trial, font, size) <= width or not cur:
            cur = trial
        else:
            lines.append(cur)
            cur = word
    return lines + [cur] if cur else lines or [""]


# --------------------------------------------------------------------------- #
# Self-check: PYTHONPATH=<repo root> PANELBUILDER_CACHE=tests/fixtures python app/exports.py
# --------------------------------------------------------------------------- #

def _pdf_streams(data: bytes) -> list[str]:
    """The page content streams, in page order, decoded."""
    import base64
    import zlib
    return [zlib.decompress(base64.a85decode(s)).decode("latin-1")
            for s in re.findall(rb"stream\r?\n(.*?)~>\s*endstream", data, re.S)]


def _pdf_pages(data: bytes) -> list[str]:
    """The strings a reader actually sees, per page, in page order.

    Not a general PDF parser: it reads the ASCII85+Flate streams and text-showing
    operators reportlab happens to emit, which is what drawString put there.
    """
    return ["\n".join(re.findall(r"\((.*?)\) Tj", d)) for d in _pdf_streams(data)]


def _pdf_text(data: bytes) -> str:
    """Every string a reader sees, whole document."""
    return "\n".join(_pdf_pages(data))


def _pdf_baselines(data: bytes) -> list[list[float]]:
    """The y every line of text was drawn at, per page: what a clipping check reads."""
    return [[float(y) for y in re.findall(r"1 0 0 1 -?[\d.]+ (-?[\d.]+) Tm", d)]
            for d in _pdf_streams(data)]


def _pdf_rects(data: bytes, op: str) -> list[tuple]:
    """Rectangles per page, by paint operator: 'f\\*' filled, 'S' stroked only.

    The warning boxes are the only rectangles this report paints, and the distinction is the
    warning: a filled bar is a danger and an outline is a pair nothing checked.
    """
    pat = r"n (-?[\d.]+) (-?[\d.]+) (-?[\d.]+) (-?[\d.]+) re " + op
    return [tuple(tuple(float(g) for g in r) for r in re.findall(pat, d))
            for d in _pdf_streams(data)]


def _pdf_stems(data: bytes) -> list[float]:
    """Heights of the locus figure's lollipop stems, read back off the page.

    Stems are the only vertical segments rising from the axis (ticks drop below it, the
    variant's dashed line starts 5pt lower), so they are the up-segments sharing the
    commonest bottom edge. Never assumes where axis_y is, so re-layout cannot break it.
    """
    import collections
    segs = []
    for d in _pdf_streams(data):
        for x1, y1, x2, y2 in re.findall(r"([\d.]+) ([\d.]+) m\s+([\d.]+) ([\d.]+) l", d):
            if abs(float(x1) - float(x2)) < 0.01 and float(y2) > float(y1):
                segs.append((round(float(y1), 2), float(y2) - float(y1)))
    if not segs:
        return []
    axis_y = collections.Counter(y for y, _ in segs).most_common(1)[0][0]
    return [h for y, h in segs if y == axis_y]


def _pdf_mark_points(stream: str) -> list[tuple[float, float]]:
    """Page-space points of the monogram, read off one page's content stream.

    The mark is the only many-sided polygon the report draws: every other straight line on
    the page is a 2-point segment and every other curve is a bezier, so a moveTo carrying
    two or more lineTos identifies it without assuming where it was placed.
    """
    for x, y, rest in re.findall(r"(-?[\d.]+) (-?[\d.]+) m((?:\s+-?[\d.]+ -?[\d.]+ l)+)", stream):
        pts = [(float(x), float(y))] + [(float(a), float(b)) for a, b in
                                        re.findall(r"(-?[\d.]+) (-?[\d.]+) l", rest)]
        if len(pts) >= 3:
            return pts
    return []


if __name__ == "__main__":
    import os
    import sys

    os.environ.setdefault("PANELBUILDER_CACHE",
                          str(pathlib.Path(__file__).resolve().parent.parent / "tests" / "fixtures"))
    r = pb.build("NM_000352.6(ABCC8):c.3989-9G>A")
    assert r.variant.rsid == "rs151344623", r.variant.rsid

    blobs = {ext: fn(r) for ext, fn in
             [("csv", to_csv), ("json", to_json), ("xlsx", to_xlsx), ("pdf", to_pdf)]}
    for ext, data in blobs.items():
        name = FILENAME(r, ext)
        assert data, f"{ext} is empty"
        pathlib.Path("/tmp", name).write_bytes(data)
        print(f"  /tmp/{name:52} {len(data):>9,} bytes")

    # R8: the disclaimer must survive verbatim into the text formats.
    for ext in ("csv", "json"):
        assert pb.DISCLAIMER.encode() in blobs[ext], f"disclaimer missing from {ext}"
    assert blobs["xlsx"][:2] == b"PK", "xlsx magic bytes"
    assert blobs["pdf"][:4] == b"%PDF", "pdf magic bytes"

    # R8 in the binaries too: xlsx zip entries are deflated, so read it back.
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(blobs["xlsx"]))
    assert wb.sheetnames == ["Recommended panel", "All candidates",
                             "Variant + provenance", "Using these markers"], wb.sheetnames
    for ws in wb:
        assert ws["A1"].value == pb.DISCLAIMER, f"disclaimer missing from sheet {ws.title!r}"
    assert wb["All candidates"].max_row == len(r.candidates) + 3, "candidate row count"

    # R4/R6/R7 spot checks on the CSV header + table.
    head, table = blobs["csv"].decode().split("rsid,chrom,pos_grch38", 1)
    for must in ("2pq", "prior", "GRCh38", "transcript sense", "minus strand",
                 pb.LAYER_B_STEPS[0], "C>T"):
        assert must in head, f"CSV header missing {must!r}"
    assert "het_2pq_prior_global" in ",".join(CSV_COLUMNS)
    assert table.count("\n") >= len(r.candidates)

    js = json.loads(blobs["json"])
    assert js["disclaimer"] == pb.DISCLAIMER
    assert js["layer_b_steps"] == pb.LAYER_B_STEPS
    assert js["provenance"]["build"] == "GRCh38"
    # Pinned to the result, not a literal: the candidate count moves legitimately whenever
    # upstream QC filters change.
    assert len(js["candidates"]) == len(r.candidates) > 1000
    assert sum(m["in_recommended_panel"] for m in js["candidates"]) == len(r.recommended)
    assert js["variant"]["pos_grch38"] == 17397055
    assert js["rarity"]["population_LD_usable"] is False

    # --- branches the golden case does not hit -------------------------------
    import copy

    def no_primers(res):
        """A copy of res carrying no primer at all: what a server without primer3 builds.

        Whether THIS server has the optional dependency is not the golden report's business,
        so every check that pins the report's shape is pinned against this.
        """
        out = copy.deepcopy(res)
        for mk in out.candidates:
            mk.primer = None
        return out

    # 1. Approximate genetic map => every format must say so next to the cM column.
    approx = copy.deepcopy(r)
    approx.candidates[0].map_approx = True
    assert "APPROXIMATE" in _cm_note(approx)
    assert b"APPROXIMATE" in to_csv(approx)
    assert "APPROX" in json.loads(to_json(approx))["provenance"]["genetic_map_note"].upper()
    wb2 = load_workbook(io.BytesIO(to_xlsx(approx)))
    assert any("APPROX" in str(c.value) for c in wb2["All candidates"][3]), "xlsx cM note"
    assert to_pdf(approx)[:4] == b"%PDF"

    # 2. No rsID (bare-HGVS resolve) => filename falls back to the sanitised query.
    anon = copy.deepcopy(r)
    anon.variant.rsid = None
    # Pin the stem, not the calendar: the date comes from the result's own provenance.
    fn = FILENAME(anon, "csv")
    assert fn.startswith("originmarker_NM_000352_6_ABCC8_c_3989_9G_A_GRCh38_"), fn
    assert fn.endswith(".csv") and fn[-14:-4] == r.provenance["queried_utc"][:10], fn
    assert to_pdf(anon)[:4] == b"%PDF"      # figure labels the variant without an rsID

    # 3. A panel too tall for one page must paginate, not overflow.
    big = copy.deepcopy(r)
    big.recommended = (big.recommended * 6)[:110]
    pdf = to_pdf(big)
    assert pdf.count(b"/Type /Page\n") >= 2 or pdf.count(b"/Type/Page") >= 2, "expected 2+ pages"
    assert len(pdf) > len(blobs["pdf"])

    # 4. The PDF says what the MODEL says about which side a marker is on, and never
    # names a chromosome arm.
    pdf_text = _pdf_text(blobs["pdf"])
    assert "rs757110" in pdf_text and "Recommended panel" in pdf_text, "pdf text extraction"
    for m in r.recommended:
        assert m.side in pdf_text, f"PDF does not render m.side {m.side!r} for {m.rsid}"
    for fmt in ("pdf", "csv"):
        body = pdf_text if fmt == "pdf" else blobs["csv"].decode()
        arm = re.findall(r"\b(?:tel|cen)\b", body)
        assert not arm, f"{fmt} still claims a chromosome arm: {arm[:4]}"

    # ...and renders that field rather than deriving it. The engine sets side from the sign
    # of dist, so only a marker whose side CONTRADICTS its sign can tell the two apart.
    q_arm = copy.deepcopy(r)
    q_arm.recommended[0].side = "SIDE_FROM_MODEL"
    assert q_arm.recommended[0].dist < 0
    assert "SIDE_FROM_MODEL" in _pdf_text(to_pdf(q_arm)), "PDF recomputes side, not renders it"
    assert "SIDE_FROM_MODEL" in to_csv(q_arm).decode()

    # 5. Every quantity that decides in_recommended_panel is a column in every export.
    # With an ancestry the sort keys on that population's 2pq, so it must be readable
    # alongside the global one.
    eas = pb.build(pb.StructuredQuery(variant="NM_000352.6(ABCC8):c.3989-9G>A", ancestry="EAS"))
    assert _rank_pop(eas) == "EAS"
    m0 = max(eas.recommended, key=lambda m: abs((_rank_het(m, "EAS") or 0) - m.het))
    assert _rank_het(m0, "EAS") is not None and abs(_rank_het(m0, "EAS") - m0.het) > 0.05, \
        "need a marker whose EAS 2pq differs from global, or this check proves nothing"
    eas_pdf = _pdf_text(to_pdf(eas))
    assert "2pq EAS" in eas_pdf, "PDF must head the deciding column with the ancestry"
    assert f"{_rank_het(m0, 'EAS'):.3f}" in eas_pdf, "PDF must print the EAS 2pq it ranked on"
    # Read the deciding column back BY NAME and check the VALUE under it, in each format:
    # presence-only and length-only asserts both pass while the artifact is wrong, since
    # _columns()/_row() are parallel lists that can drift out of order at equal length.
    eas_csv = to_csv(eas).decode().split("\n")
    hdr_i = next(i for i, l in enumerate(eas_csv) if l.startswith("rsid,"))   # find it, don't count back
    eas_cols = eas_csv[hdr_i].split(",")
    eas_xl = load_workbook(io.BytesIO(to_xlsx(eas)))["All candidates"]
    xl_hdr = [c.value for c in eas_xl[3]]
    # Keyed by variant_id, NOT rsid: multi-allelic sites share an rsID, so an rsid-keyed
    # dict drops every alt but the last and compares the wrong row.
    eas_json = {c["variant_id"]: c for c in json.loads(to_json(eas))["candidates"]}
    for fmt, cols in (("csv", eas_cols), ("xlsx", xl_hdr)):
        assert "het_2pq_prior_EAS" in cols, f"{fmt} has no het_2pq_prior_EAS column"
    for m, csv_line, xl_row in zip(eas.candidates, eas_csv[hdr_i + 1:],
                                   eas_xl.iter_rows(min_row=4, values_only=True)):
        want = _rank_het(m, "EAS")
        by_name = dict(zip(_columns("EAS"), _row(m, set(), "EAS")))
        assert by_name["het_2pq_prior_EAS"] == want, "_columns/_row are out of order"
        assert by_name["rsid"] == m.rsid, "_columns/_row are out of order"
        got_csv = csv_line.split(",")[eas_cols.index("het_2pq_prior_EAS")]
        assert got_csv == ("" if want is None else str(want)), \
            f"csv: het_2pq_prior_EAS is {got_csv!r} under its own header, want {want!r}"
        assert xl_row[xl_hdr.index("het_2pq_prior_EAS")] == want, "xlsx: wrong value under header"
        assert eas_json[m.variant_id]["het_2pq_prior_EAS"] == want, \
            f"json: het_2pq_prior_EAS is {eas_json[m.variant_id]['het_2pq_prior_EAS']!r}, want {want!r}"
    assert len(_columns("EAS")) == len(_row(m0, set(), "EAS")) == len(CSV_COLUMNS) + 1
    # No ancestry: the decider is the global 2pq, which is on the page under its own name.
    assert "2pq global" in pdf_text and "2pq prior" not in pdf_text

    # _decider mirrors panelbuilder._rank_key across a seam: pin it to the engine's actual
    # key so a change of ranking quantity fails here, not in a shipped export.
    for res, a in ((r, None), (eas, "EAS")):
        for m in res.recommended:
            engine_het = -pb._rank_key(a)(m)[0]
            # Tolerance covers _het's round-to-4dp and nothing more. NOT 5e-5: the engine
            # keeps 2p(1-p) unrounded, and ordinary MAFs land on that exact boundary just
            # over a strict < 5e-5. A change of ranking quantity is orders bigger.
            assert abs(engine_het - _decider(m, a)) < 1e-4, \
                f"exports and the engine disagree on what ranked {m.rsid}: " \
                f"{_decider(m, a)} vs {engine_het}"

    # 6. Seam contract: exports render provenance["ranking_key"] verbatim and never
    # restate the ranking basis in their own words.
    keyed = copy.deepcopy(r)
    keyed.provenance["ranking_key"] = "SENTINEL-KEY 2pq in XYZ"
    assert _ranking_key(keyed) == "SENTINEL-KEY 2pq in XYZ"
    assert "SENTINEL-KEY 2pq in XYZ" in _pdf_text(to_pdf(keyed))
    assert "SENTINEL-KEY 2pq in XYZ" in to_csv(keyed).decode()
    assert json.loads(to_json(keyed))["provenance"]["ranking_key"] == "SENTINEL-KEY 2pq in XYZ"
    for body in (pdf_text, blobs["csv"].decode()):
        assert "none (global 2pq prior)" not in body, "exports restating the ranking basis"
    if not r.provenance.get("ranking_key"):
        print("  NOTE: panelbuilder sets no provenance['ranking_key'] yet; exports render "
              "the marked fallback.", file=sys.stderr)

    # 7. The FIGURE draws the number it ranked on: every check above reads text, none the
    # geometry. Scale-free on purpose, asserting heights are PROPORTIONAL to _decider
    # rather than equal to a layout constant, so re-sizing the figure does not fail this.
    dec = sorted(_decider(m, "EAS") for m in eas.recommended)
    stems = sorted(_pdf_stems(to_pdf(eas)))
    assert len(stems) == len(dec), f"{len(stems)} lollipops for {len(dec)} markers"
    ratios = [h / d for h, d in zip(stems, dec)]
    assert max(ratios) - min(ratios) < 1.0, \
        f"lollipop heights are not proportional to the 2pq that ranked them: {ratios[:3]}"

    # 8. The no-frequency fallback: "-" in the column, global 2pq in the ranking. The
    # marker is synthesised because no fixture in this repo reaches this branch: ABCC8 has
    # a gnomAD frequency for every recommended marker in all 8 populations.
    gap = copy.deepcopy(eas)
    m_gap = gap.recommended[0]
    del m_gap.per_pop_maf["EAS"]
    assert _rank_het(m_gap, "EAS") is None
    assert _decider(m_gap, "EAS") == m_gap.het, "fallback must be the global 2pq, as the engine does"
    assert -pb._rank_key("EAS")(m_gap)[0] == m_gap.het, "engine disagrees about the fallback"
    gap_row = [l for l in _pdf_text(to_pdf(gap)).split("\n")]
    i = gap_row.index(m_gap.rsid)
    assert gap_row[i + 6] == "-", f"no-frequency marker must print '-', printed {gap_row[i + 6]!r}"
    assert "no EAS frequency" in _pdf_text(to_pdf(gap)), "PDF must footnote what the dash means"
    assert _row(m_gap, set(), "EAS")[13] is None, "csv/xlsx cell must be empty, never a number"

    # 9. A model-chosen panel must SAY SO in all four formats. Synthesised: the golden
    # ABCC8 case was named by the user, so no fixture reaches this branch.
    chosen = copy.deepcopy(r)
    chosen.provenance["nl_model"] = "claude-test-model-1"
    chosen.provenance["nl_text"] = "markers near the SENTINEL-PROSE splice mutation"
    for fmt, body in (("csv", to_csv(chosen).decode()),
                      ("json", to_json(chosen).decode()),
                      ("pdf", _pdf_text(to_pdf(chosen))),
                      ("xlsx", "\n".join(
                          str(c.value) for ws in load_workbook(io.BytesIO(to_xlsx(chosen)))
                          for row in ws.iter_rows() for c in row if c.value is not None))):
        assert "language model" in body, f"{fmt} does not say a model chose the variant"
        assert "claude-test-model-1" in body, f"{fmt} does not name the model"
        assert "SENTINEL-PROSE" in body, f"{fmt} does not quote the text the model read"
    # The CSV caveat trap: pandas.read_csv(comment='#') drops the whole header block, so
    # the fact has to be in the TABLE as well or that reader never sees it.
    stripped = "\n".join(l for l in to_csv(chosen).decode().split("\n")
                         if not l.startswith("#"))
    assert "variant_chosen_by_language_model" in stripped and "claude-test-model-1" in stripped, \
        "a reader stripping '#' comments cannot see that a model chose the variant"
    assert json.loads(to_json(chosen))["provenance"]["nl_model"] == "claude-test-model-1"

    # ...and a panel the user named stays SILENT: no caveat, no "none", no column.
    assert _nl_caveat(r) is None
    assert r.provenance["nl_model"] is None, "the golden fixture must not be model-chosen"
    for fmt, body in (("csv", blobs["csv"].decode()), ("json", blobs["json"].decode()),
                      ("pdf", pdf_text)):
        assert "language model" not in body, f"{fmt} caveats a panel the user typed"
        assert "nl_model: none" not in body.lower(), f"{fmt} renders a None nl_model"
    assert "variant_chosen_by_language_model" not in blobs["csv"].decode()

    # 10. The mark. web/public/favicon.svg is its one definition, so every check here is
    # that the artifact follows the FILE, never that it matches a literal written here.
    from reportlab.lib.pagesizes import letter

    mark, real_svg = _mark(), list(MARK_SVG)
    assert mark, f"favicon.svg not found at any of {[str(p) for p in MARK_SVG]}"
    pts = _pdf_mark_points(_pdf_streams(blobs["pdf"])[0])
    assert len(pts) == sum(len(p) for p in mark["polys"]), \
        f"the PDF drew {len(pts)} points; favicon.svg has {sum(len(p) for p in mark['polys'])}"
    # A floor, not the geometry: the M is a real outlined glyph and no redraw of one runs
    # to this many points, so this fails if the mark is ever swapped for an approximation.
    assert len(pts) > 20, f"the PDF's mark is down to {len(pts)} points: a redraw, not a glyph"
    # The masthead sits above the report, not through it: it can only displace what it sits
    # on top of, which is the card, so the card is what this pins.
    page1 = "\n".join(re.findall(r"\((.*?)\) Tj", _pdf_streams(blobs["pdf"])[0]))
    assert page1.startswith("OriginMarker - candidate linkage markers for PGT-M\n"
                            "rs151344623 ABCC8"), "the masthead displaced the variant card"
    for lbl in (r"Genomic \(VCF\):", r"Coverage \(R5\):"):     # first and last card rows
        assert lbl in page1, f"the masthead pushed {lbl!r} off page 1"
    # Length is a canary, not a target: any growth trips this and gets looked at. Measured
    # without primers, since how many blocks this server can design is not the report's shape.
    assert len(_pdf_streams(to_pdf(no_primers(r)))) == 2, "the golden report changed length"
    assert min(p[1] for p in pts) > letter[1] - 72, "the mark left the top inch of the page"

    # Page 1 only. A mark on every page is noise, and the multipage branch is where a
    # header leaks into the running pages.
    big_pages = _pdf_streams(pdf)
    assert len(big_pages) > 1 and [i for i, s in enumerate(big_pages) if _pdf_mark_points(s)] \
        == [0], "the mark must be a page-1 masthead, not a running header"

    # A mark it cannot draw exactly is refused, not approximated: a curve flattened to a
    # chord, or a relative command read as absolute, is a wrong mark drawn confidently.
    for bad in ("M0 0C10 10 20 20 30 30Z", "m0 0l10 10Z", "M0 0Q5 5 10 0Z"):
        try:
            _svg_polys(bad)
            assert False, f"_svg_polys redrew {bad!r} as straight lines"
        except ValueError:
            pass

    # The PDF's mark IS favicon.svg rather than a copy of its numbers: give the loader a
    # different file and the drawn geometry must follow it, in page space. A right isoceles
    # triangle on the viewBox pins scale, aspect and the SVG-to-PDF y flip at once; the
    # real mark's own path cannot, since only the file knows what it should look like.
    probe = pathlib.Path("/tmp/originmarker_probe_favicon.svg")
    probe.write_bytes(b'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 100 100">'
                      b'<circle cx="50" cy="50" r="38" fill="none" stroke="#2e6da4" '
                      b'stroke-width="10"/><path d="M0 0L100 0L100 100Z" fill="#337ab7"/></svg>')
    MARK_SVG = [probe]
    assert _mark()["polys"] == [[(0.0, 0.0), (100.0, 0.0), (100.0, 100.0)]], "path parse"
    (x0, y0), (x1, y1), (x2, y2) = _pdf_mark_points(_pdf_streams(to_pdf(r))[0])
    size = x1 - x0
    assert size > 0 and abs(y1 - y0) < 0.01, "the mark's top edge is not horizontal"
    assert abs(x2 - x1) < 0.01 and abs((y1 - y2) - size) < 0.01, \
        f"the mark is not square and y-flipped onto the page: {(x0, y0), (x1, y1), (x2, y2)}"
    MARK_SVG = real_svg
    assert _mark()["polys"] == mark["polys"], "the probe leaked into the real mark"

    # A missing SVG costs the reader a logo, never the report: the data outranks the brand.
    MARK_SVG = [pathlib.Path("/nonexistent/favicon.svg")]
    assert _mark() is None and _mark_png(18) is None
    assert _pdf_text(to_pdf(r)).startswith("OriginMarker - candidate"), \
        "with no SVG the PDF must still title itself"
    assert to_xlsx(r)[:2] == b"PK", "with no SVG the workbook must still build"
    MARK_SVG = real_svg

    # The XLSX carries the mark on its first sheet only, and the bytes it carries are the
    # ones favicon.svg drew. Read through the zip, NOT through load_workbook: openpyxl's
    # reader reports one image for a workbook that has one on every sheet, so a check
    # written against it passes while the file is wrong.
    import zipfile
    z = zipfile.ZipFile(io.BytesIO(blobs["xlsx"]))
    assert z.read("xl/media/image1.png") == _mark_png(54), "the workbook's mark is not the SVG's"
    assert [n for n in z.namelist() if re.fullmatch(r"xl/drawings/drawing\d+\.xml", n)] \
        == ["xl/drawings/drawing1.xml"], "the mark repeats past the first sheet"
    assert b"drawing1.xml" in z.read("xl/worksheets/_rels/sheet1.xml.rels"), \
        "the mark is not on sheet 1"
    wb3 = load_workbook(io.BytesIO(blobs["xlsx"]))
    assert wb3.sheetnames[0] == "Recommended panel"          # ...which is sheet1.xml
    assert wb3["Recommended panel"]["A1"].value == pb.DISCLAIMER, "R8 text must be untouched"
    assert wb3["Recommended panel"]["A1"].alignment.indent >= 2, "the mark covers the disclaimer"

    # 11. The star. The engine decides it and words it; exports read both and re-decide
    # neither, so the seam is pinned to pb rather than to a copy of its text written here.
    NOTE = pb.FLANKING_CRITERIA["note"]
    assert _flanking(r)["field"] == STAR_FIELD and _flanking(r)["note"] == NOTE, \
        "exports and the engine disagree about the star's field or its words"
    star_n = sum(_starred(m) for m in r.recommended)
    assert 0 < star_n < len(r.recommended), \
        "the golden panel must star SOME markers and not all, or this proves nothing"

    star_pdf = _pdf_text(blobs["pdf"])
    star_xl = load_workbook(io.BytesIO(blobs["xlsx"]))
    # _wrap breaks a paragraph across drawString calls and a content stream escapes its
    # parens: undo both, so what is compared is the sentence the reader sees.
    star_flat = " ".join(star_pdf.split()).replace("\\(", "(").replace("\\)", ")")
    for fmt, body in (("csv", blobs["csv"].decode()), ("json", blobs["json"].decode()),
                      ("pdf", star_flat),
                      ("xlsx", "\n".join(str(c.value) for ws in star_xl
                                         for row in ws.iter_rows()
                                         for c in row if c.value is not None))):
        # Verbatim, from the engine. An export's own retelling of the criteria is what
        # drifts, and the star is the half that would still be printed after it did.
        for para in NOTE:
            want = para.lstrip("★ ") if fmt == "pdf" else para
            assert " ".join(want.split()) in body, \
                f"{fmt} does not render the engine's own words for the star"

    # A real column, and it survives the reader who strips the header block:
    # pandas.read_csv(comment='#') deletes every word of the legend above it.
    bare = [l for l in blobs["csv"].decode().split("\n") if not l.startswith("#") and l]
    star_i = bare[0].split(",").index(STAR_FIELD)
    # Three states, not two. The predicate runs over the shortlist only, so a candidate it
    # never judged must print EMPTY: a False there is a failing verdict on a marker nobody
    # assessed, and several unjudged ones sit nearer the variant than the starred ones.
    for m, line in zip(r.candidates, bare[1:]):
        cell = line.split(",")[star_i]
        want = "" if _star_cell(m) is None else str(_star_cell(m))
        assert cell == want, \
            f"csv: wrong {STAR_FIELD} under its own header for {m.rsid}: {cell!r} != {want!r}"
    judged = [m for m in r.candidates if _star_cell(m) is not None]
    assert len(judged) == len(r.recommended), "the verdict is the shortlist's, and only its"
    assert sum(1 for l in bare[1:] if l.split(",")[star_i] == "") \
        == len(r.candidates) - len(r.recommended), "unjudged candidates must print empty"
    assert sum(l.split(",")[star_i] == "True" for l in bare[1:]) == star_n, "csv star count"
    xl_rec = star_xl["Recommended panel"]
    xl_c = [c.value for c in xl_rec[3]].index(STAR_FIELD)
    assert [row[xl_c] for row in xl_rec.iter_rows(min_row=4, values_only=True)] \
        == [_starred(m) for m in r.recommended], "xlsx: wrong value under the star header"
    sj = {c["variant_id"]: c for c in json.loads(blobs["json"])["candidates"]}
    assert [sj[m.variant_id][STAR_FIELD] for m in r.recommended] \
        == [_starred(m) for m in r.recommended], "json: wrong star per marker"

    # The PDF prints a SHAPE, in a standard font that has one: no text font carries U+2605,
    # and a colour is the first thing the printer this page is filed off throws away.
    # Read back the CODE ZapfDingbats encoded U+2605 to, never the character handed in:
    # passing that code instead of the character draws a different dingbat just as quietly.
    STAR_CODE = "H"
    assert b"/ZapfDingbats" in blobs["pdf"], "the PDF does not carry the star's font"
    star_rows = star_pdf.split("\n")
    for m in r.recommended:
        i = star_rows.index(m.rsid)
        assert (star_rows[i - 1] == STAR_CODE) == _starred(m), \
            f"PDF star on {m.rsid} is {star_rows[i - 1]!r}, the engine says {_starred(m)}"
    # Per side, against ESHRE's per-side minimum: a total answers the wrong question.
    for k in ("lower_flanking_count", "higher_flanking_count"):
        assert f"{r.coverage[k]} {k.split('_')[0]}-coordinate" in star_flat, \
            f"the PDF does not read out coverage[{k!r}]"

    # An unstarred panel says NOTHING: no legend, no glyph, no restated rule. Synthesised,
    # because the golden panel stars markers and no fixture reaches a panel that stars none.
    none_starred = copy.deepcopy(r)
    for m in none_starred.candidates:
        setattr(m, STAR_FIELD, False)
    assert _star_legend(none_starred, none_starred.candidates) == []
    none_csv = to_csv(none_starred).decode()
    assert "hropen" not in none_csv and "informativity" not in none_csv, \
        "a legend for a star the panel does not show"
    assert STAR_CODE not in _pdf_text(to_pdf(none_starred)).split("\n"), \
        "an empty star column on the PDF"
    assert STAR_FIELD in none_csv, "the column is data and must stay whatever the verdict"

    # The words come from the RESULT's stamped rule, not from whatever pb holds at export
    # time: a re-worded rule must not silently re-label a panel built under the old one.
    reworded = copy.deepcopy(r)
    reworded.provenance["flanking_criteria"] = dict(_flanking(r), note=["SENTINEL-NOTE ★ x"])
    assert "SENTINEL-NOTE" in to_csv(reworded).decode()
    assert "SENTINEL-NOTE" in " ".join(_pdf_text(to_pdf(reworded)).split())
    # ...and a build that stamps no rule prints the marked fallback rather than inventing
    # a meaning for a glyph it is still drawing.
    unnamed = copy.deepcopy(r)
    del unnamed.provenance["flanking_criteria"]
    assert "not reported by this build" in _star_legend(unnamed, unnamed.recommended)[0]
    for k in ("lower_flanking_count", "higher_flanking_count"):
        del unnamed.coverage[k]
    assert "not reported by this build" in _star_count_text(unnamed), \
        "an uncounted side must say so, never read as zero"

    # An allele frequency is prose on the PDF and a data cell in the CSV: the PDF renders
    # it via fmt_af, the CSV column keeps the exact float.
    pdf_txt = _pdf_text(to_pdf(r))
    assert repr(r.rarity.gnomad_af_genome) not in pdf_txt, "float repr leaked into the PDF"
    assert pb.fmt_af(r.rarity.gnomad_af_genome) in pdf_txt, "PDF should render the AF via fmt_af"
    af_col = _columns(None).index("af")
    assert _row(r.recommended[0], set())[af_col] == r.recommended[0].af, \
        "the CSV af column must stay an exact float, never a formatted string"

    # 12. Primers. Synthesised, because primer design is optional and is not in the default
    # image: the golden fixture may carry no pair at all, and every branch below still has to
    # be proven. A panel with no primer is the OTHER case, and it is `r` itself.
    import primers

    # A panel with no primer says NOTHING about primers: no column, no block, no "0 designed".
    # This is the default server (primer3 is optional and absent from the default image), and
    # a report that describes a design failure that never happened is describing nothing.
    nop = no_primers(r)
    nop_pdf = to_pdf(nop)
    for fmt, body in (("csv", to_csv(nop).decode()), ("pdf", _pdf_text(nop_pdf)),
                      ("xlsx", "\n".join(
                          str(c.value) for ws in load_workbook(io.BytesIO(to_xlsx(nop)))
                          for row in ws.iter_rows() for c in row if c.value is not None))):
        assert "primer" not in body.lower(), \
            f"{fmt} talks about primers for a build that designed none"
    # JSON keeps the field, because Marker.primer is one and a machine reader wants a stable
    # key: null is the honest rendering of a design that never ran. The PROSE stays away.
    nop_js = json.loads(to_json(nop))
    assert all(cd["primer"] is None for cd in nop_js["candidates"]), \
        "a panel with no design must carry a null primer, never a record"
    assert not [k for k in nop_js["provenance"] if "primer" in k], \
        "json explains primers to a build that designed none"

    def dna(i, n):
        """A distinct, plausible oligo per marker: the index in base 4, then filler."""
        s = ""
        while i:
            s, i = s + "ACGT"[i % 4], i // 4
        return (s + "GCATTGCAAGGCCTTAGCATCGATCGGATCCTTAGCAA")[:n]

    def oligo(i, n, pos):
        return primers.Primer(seq=dna(i, n), pos=pos, idx=i, length=n,
                              tm=68.5 + i / 10, gc=54.5)

    # Built off the primer-free panel, so this case is the same whether or not this server
    # has primer3: the engine's own designs are its module's to prove, not this one's.
    prim = no_primers(r)
    starred = [m for m in prim.recommended if _starred(m)]
    assert len(starred) >= 8, \
        f"the multipage primer case needs several starred markers, this panel has {len(starred)}"
    # Real primers.PrimerResult objects, never a stand-in shape: this is the seam exports
    # read, so a renamed field must fail here rather than ship an empty primer block.
    for i, m in enumerate(starred):
        kw = dict(fwd=oligo(2 * i + 2, 26, m.pos - 190), rev=oligo(2 * i + 3, 24, m.pos + 160),
                  product_size=380 + i, product_start=m.pos - 190,
                  masked=tuple(primers.MaskSite(pos=m.pos - 300 + j, ref="A", alt="G",
                                                maf=0.02) for j in range(4 + i)),
                  mask_note=primers.Note(
                      code="mask",
                      short=f"SENTINEL-SHORT-mask-{i}: the one-line form.",
                      long=f"SENTINEL-MASK-{i}: gnomAD SNPs and indels at MAF >= 1% in this "
                           f"window sit under neither primer."),
                  warnings=(primers.NOT_CHECKED_WARNING,))
        # Both lengths carry a distinguishable sentinel, because print must take the long one
        # and only the long one: the short form defers to a docs link, and a filed page has
        # nothing to click. A PDF quietly rendering `short` would still look like prose.
        def note(code, long):
            return primers.Note(code=code, long=long,
                                short=f"SENTINEL-SHORT-{code}-{i}: the one-line form.")
        # Five cases, because there are five: the four ispcr verdicts and a design that
        # failed. not_checked is the DEFAULT and the one every real build ships today, so it
        # is the one that must not be missing from the check.
        if i % 5 == 0:
            kw.update(insilico_pcr=ispcr.DANGER,
                      warnings=(note("danger",
                                     f"SENTINEL-DANGER-{i}: UCSC In-Silico PCR found {2 + i} "
                                     f"products for this pair in hg38, not one. A pair that "
                                     f"amplifies more than one locus cannot be genotyped. Do "
                                     f"not order this pair without redesigning it."),
                                note("caveat", ispcr.CAVEAT)))
        elif i % 5 == 1:
            kw.update(insilico_pcr=ispcr.NOT_CHECKED)      # the engine's own default words
        elif i % 5 == 2:
            kw.update(insilico_pcr=ispcr.UNKNOWN,
                      warnings=(note("unknown",
                                     f"SENTINEL-UNKNOWN-{i}: the answer from UCSC could not "
                                     f"be read, so this pair is not verified. Retry later."),))
        elif i % 5 == 3:
            kw.update(insilico_pcr=ispcr.ONE_PRODUCT,
                      warnings=(note("pass",
                                     f"SENTINEL-PASS-{i}: UCSC In-Silico PCR: one product, "
                                     f"chr{m.chrom}, {380 + i}bp, as designed."),
                                note("caveat", ispcr.CAVEAT)))
        else:
            kw = dict(mask_note=kw["mask_note"], masked=kw["masked"],
                      error=f"SENTINEL-FAIL-{i}: no {primers.DEFAULTS.min_size}-"
                            f"{primers.DEFAULTS.max_size} bp primer reaches "
                            f"{primers.DEFAULTS.min_tm:.0f} C in this window at "
                            f"{primers.DEFAULTS.min_gc:.0f}-{primers.DEFAULTS.max_gc:.0f}% "
                            f"GC; the window is 30% GC. Nothing was relaxed to force a pair.")
        m.primer = primers.PrimerResult(**kw)
    pm = [m for m in prim.candidates if _primer(m)]
    assert len(pm) == len(starred) and _has_primers(prim)
    n_danger = sum(_pcr_state(_primer(m)) == "danger" for m in pm)
    n_box = sum(1 for m in pm if not _primer(m).ok
                or _pcr_state(_primer(m)) == "unverified")
    assert n_danger and n_box and sum(_pcr_state(_primer(m)) == "pass" for m in pm), \
        "the primer case must carry a danger, an unverified pair and a clean one, or it " \
        "proves nothing about telling them apart"

    p_pdf, p_csv = to_pdf(prim), to_csv(prim).decode()
    p_json = json.loads(to_json(prim))
    p_xl = load_workbook(io.BytesIO(to_xlsx(prim)))
    pages = _pdf_pages(p_pdf)
    p_flat = " ".join(_pdf_text(p_pdf).split()).replace("\\(", "(").replace("\\)", ")")

    # The whole panel of blocks pushes the report past one page, which is where a masthead
    # leaks, a footer gets overrun and a block gets torn in half.
    assert len(pages) > len(_pdf_streams(nop_pdf)), "primer blocks did not lengthen the PDF"
    assert [i for i, s in enumerate(_pdf_streams(p_pdf)) if _pdf_mark_points(s)] == [0], \
        "the mark must stay a page-1 masthead once the primers paginate"
    for i, page in enumerate(pages):                 # R8, verbatim, on EVERY page
        flat = " ".join(page.split()).replace("\\(", "(").replace("\\)", ")")
        assert " ".join(pb.DISCLAIMER.split()) in flat, f"disclaimer missing from page {i + 1}"
    # Nothing clips: every line of text sits inside the page, above the footer's own band.
    for i, ys in enumerate(_pdf_baselines(p_pdf)):
        assert ys and min(ys) >= 26 and max(ys) <= letter[1] - 26, \
            f"page {i + 1} draws text outside the page: {min(ys)} to {max(ys)}"

    # A block is never split across a page: whatever page carries a pair's sequences carries
    # its warning too. A warning left behind on the previous page is a primer with no warning.
    for m in pm:
        p = _primer(m)
        key = p.error or _primer_warnings(p)[0]
        home = [i for i, pg in enumerate(pages) if m.rsid in pg.split("\n")]
        assert home, f"{m.rsid} has no primer block on any page"
        pg = " ".join(pages[home[-1]].split())
        for part in ([f"FWD 5'-{p.fwd.seq}-3'", f"REV 5'-{p.rev.seq}-3'"] if p.ok else []) \
                + [" ".join(key.split()[:6])]:
            assert " ".join(part.split()) in pg, \
                f"{m.rsid}: the block is split across a page, {part[:32]!r} left behind"

    # WARNINGS IN BLACK AND WHITE. A danger is reversed out of a filled bar and an unchecked
    # pair is boxed: read back the paint operators, since a hue is the first thing the printer
    # this page is filed off throws away.
    assert sum(len(pg) for pg in _pdf_rects(p_pdf, r"f\*")) == n_danger, \
        "a DANGEROUS pair without its filled warning bar"
    assert sum(len(pg) for pg in _pdf_rects(p_pdf, "S")) == n_box, \
        "an unchecked or undesigned pair without its box"
    assert not _pdf_rects(nop_pdf, r"f\*")[0], "the panel with no primers paints a bar"
    # Each verdict heads its own box, once per pair. Counted off the head of a drawn line,
    # since the warning prose underneath quotes these words back.
    p_lines = [ln.replace("\\(", "(").replace("\\)", ")")
               for ln in _pdf_text(p_pdf).split("\n")]
    for kind, head in (("danger", "DANGER"), ("unverified", "NOT VERIFIED"),
                       ("pass", "VERIFIED CLEAN (in silico)")):
        n = sum(_pcr_state(_primer(m)) == kind and _primer(m).ok for m in pm)
        got = sum(1 for ln in p_lines if ln.startswith(head + ":"))
        assert got == n, f"{got} {head!r} heads on the page for {n} {kind} pairs"
    assert "NOT VERIFIED" in p_flat and "VERIFIED CLEAN (in silico)" in p_flat, \
        "the unverified and the verified states must not read as the same verdict"
    assert sum(1 for ln in p_lines if ln.startswith("NO PRIMER DESIGNED:")) \
        == sum(1 for m in pm if not _primer(m).ok)

    # NEVER HIDE A PRIMER: the dangerous ones are on the page in full, sequence and all, and
    # every word the engine warned with is on the page beside them.
    for m in pm:
        p = _primer(m)
        for part in ([p.fwd.seq, p.rev.seq, _mask_note(p)] if p.ok else [p.error]):
            assert " ".join(str(part).split()) in p_flat, f"{m.rsid}: {str(part)[:24]!r} hidden"
        for w in (_primer_warnings(p) if p.ok else []):
            assert " ".join(w.split()) in p_flat, f"{m.rsid}: dropped a warning: {w[:40]!r}"

    # The same facts as DATA, and they survive the reader who strips the header block: the
    # sequences and the verdict are columns, never prose.
    # csv.reader, not split(","): a warning is prose and carries commas, and the reader who
    # splits on them is reading a different value under the same header.
    p_bare = [row for row in csv.reader(io.StringIO(p_csv))
              if row and not row[0].startswith("#")]
    p_cols = p_bare[0]
    for col in PRIMER_COLUMNS:
        assert col in p_cols, f"csv has no {col} column"
    p_rows = {row[0]: row for row in p_bare[1:]}
    p_xl_hdr = [c.value for c in p_xl["All candidates"][3]]
    p_xl_rows = {row[0]: row for row in
                 p_xl["All candidates"].iter_rows(min_row=4, values_only=True)}
    p_js = {c["variant_id"]: c for c in p_json["candidates"]}
    for m in prim.candidates:
        p, row = _primer(m), p_rows[m.rsid]
        want = dict(zip(PRIMER_COLUMNS, _primer_row(m)))
        assert dict(zip(_columns(None, True), _row(m, set(), None, True)))["primer_fwd"] \
            == want["primer_fwd"], "_columns/_row are out of order over the primer block"
        for col in ("primer_fwd", "primer_rev", "primer_insilico_pcr", "primer_warnings",
                    "primer_design_error"):
            v = want[col]
            assert row[p_cols.index(col)] == ("" if v is None else str(v)), \
                f"csv: wrong {col} under its own header for {m.rsid}"
            assert p_xl_rows[m.rsid][p_xl_hdr.index(col)] == v, f"xlsx: wrong {col}"
        js = p_js[m.variant_id]["primer"]
        assert (js or {}).get("primer_fwd") == want["primer_fwd"], "json: wrong sequence"
        if p is None:
            # No design attempted is not a verdict: it must print empty, never a pass.
            assert js is None and row[p_cols.index("primer_insilico_pcr")] == ""
        else:
            assert js["primer_warnings"] == _primer_warnings(p), "json drops a warning"
            assert row[p_cols.index("primer_insilico_pcr")] not in ("", "None"), \
                f"{m.rsid}: a real pair with an empty verdict cell reads as nothing wrong"
            assert row[p_cols.index("primer_warnings")], \
                f"{m.rsid}: a pair whose warnings did not survive into the CSV"
    xl_all = "\n".join(str(c.value) for ws in p_xl for row in ws.iter_rows()
                       for c in row if c.value is not None)
    for fmt, body in (("csv", p_csv), ("json", to_json(prim).decode()), ("xlsx", xl_all)):
        for token in (ispcr.NOT_CHECKED, ispcr.DANGER, ispcr.UNKNOWN):
            assert token in body, f"{fmt} drops the {token!r} verdict"
        assert "SENTINEL-DANGER-0" in body, f"{fmt} drops the danger's own words"
        # The LONG form, and never the short one. Every note carries both, and the short one
        # ends in an implicit "see the docs": printed into a file that leaves the app, it is
        # a sentence pointing at a link the reader does not have. Both are prose, so nothing
        # else here would notice the swap.
        assert "SENTINEL-SHORT-" not in body, \
            f"{fmt} prints a note's short form, which defers to a link a file cannot follow"
    assert ispcr.CAVEAT in p_csv and p_json["provenance"]["primer_insilico_pcr_caveat"] \
        == ispcr.CAVEAT, "the in-silico-only caveat must survive verbatim"

    # json writes the shortlist twice: inside `candidates` and again whole under
    # `recommended`. Both are the same marker, so both must say the same thing about it. Two
    # copies enriched by one code path is the only reason this holds.
    _by_id = {c["variant_id"]: c for c in p_json["candidates"]}
    for rmd in p_json["recommended"]:
        assert rmd == _by_id[rmd["variant_id"]], \
            f"{rmd['variant_id']}: recommended[] and candidates[] disagree about one marker"

    # A verdict this build has never heard of is NOT a pass. ispcr owns the vocabulary and
    # may grow it; an unread token rendered clean is the failure this whole module refuses.
    fake = lambda v: primers.PrimerResult(insilico_pcr=v)
    for token in ("SENTINEL-FUTURE-VERDICT", "", None, "ONE_PRODUCT", ispcr.UNKNOWN,
                  ispcr.NOT_CHECKED):
        assert _pcr_state(fake(token)) != "pass", f"{token!r} read as a pass"
    assert _pcr_state(fake(ispcr.ONE_PRODUCT)) == "pass"
    assert _pcr_state(fake(ispcr.DANGER)) == "danger"
    # ...and a pair the engine left no warning on still carries one, rather than reading as
    # a pair with nothing wrong with it.
    assert _primer_warnings(fake(ispcr.NOT_CHECKED))[0].endswith("has not been verified.")

    print(f"\nself-check OK: {len(r.candidates)} candidates, {len(r.recommended)} recommended, "
          f"4/4 formats, disclaimer verbatim in all; map-approx/no-rsid/multipage branches OK; "
          f"PDF renders m.side and the deciding 2pq, and draws it at the right height; "
          f"{len(pm)} primer blocks paginate to {len(pages)} pages, none split, "
          f"{n_danger} danger bars and {n_box} boxes drawn as shapes.", file=sys.stderr)
