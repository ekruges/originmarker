"""Pull the laboratory's own karyotype calls out of the tracking spreadsheets.

These are the manual maps offered for cross-comparison. They are free text in a Karyotype column
and a Notes column, written for people rather than for a parser, so this extracts CHROMOSOME-LEVEL
events only and refuses to guess at anything finer. A phrase like "meiotic segmental on 1" fixes
the chromosome and the class; it does not fix a breakpoint, and inventing one from it would
manufacture agreement with our own coordinates.

Output: sample, chromosome, class (whole|segmental|unspecified), verbatim source text.
"""
import glob, re, sys, csv

SRC = '/Users/ezrakruger/Downloads'
# A barcode like 52461-33 is what joins these rows to an array file.
BARCODE = re.compile(r'\b(\d{5}-\d{2})\b')
# Chromosome mentions that are safe to read: "on 1", "chr16", "16p", "trisomy 21".
CHROM = re.compile(r'\b(?:chr\.?\s*)?([12]?\d|2[0-2]|X|Y)\b(?=\s*[pq]?\b)', re.I)
SEGMENTAL = re.compile(r'segment|partial|interstitial|terminal|del\b|dup\b', re.I)
# Some notes carry a position, e.g. "meiotic segmental on 1 at 76.5Mb". Those are the rows that can
# validate a refined breakpoint; the rest fix only a chromosome.
POSMB = re.compile(r'(\d+(?:\.\d+)?)\s*(?:Mb|MB|mb)\b')
WHOLE = re.compile(r'trisomy|monosom|whole|loss of chr|gain of chr|nullisom', re.I)


def rows_from(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for name in wb.sheetnames:
        ws = wb[name]
        hdr = None
        for r in ws.iter_rows(values_only=True):
            cells = ['' if c is None else str(c) for c in r]
            if hdr is None:
                if any('karyotype' in c.lower() for c in cells):
                    hdr = {c.strip().lower(): i for i, c in enumerate(cells) if c.strip()}
                continue
            yield name, cells, hdr
    wb.close()


def main():
    out = []
    for f in glob.glob(f'{SRC}/*Zygote*.xlsx'):
        for sheet, cells, hdr in rows_from(f):
            joined = ' | '.join(cells)
            bc = BARCODE.search(joined)
            if not bc:
                continue
            ki = next((i for k, i in hdr.items() if 'karyotype' in k), None)
            ni = next((i for k, i in hdr.items() if k.startswith('note')), None)
            text = ' '.join(cells[i] for i in (ki, ni) if i is not None and i < len(cells))
            if not text.strip():
                continue
            # Only claim an event where the text names one. A karyotype like 46XX is normal.
            if re.fullmatch(r'\s*4[0-9]\s*[XY]{2}\s*', text):
                out.append((bc.group(1), '', 'euploid', text.strip(), sheet, ''))
                continue
            cls = ('segmental' if SEGMENTAL.search(text)
                   else 'whole' if WHOLE.search(text) else 'unspecified')
            chroms = {c.upper() for c in CHROM.findall(text)}
            if not chroms:
                out.append((bc.group(1), '', cls, text.strip(), sheet,
                            ';'.join(POSMB.findall(text))))
            mb = POSMB.findall(text)
            for c in sorted(chroms):
                out.append((bc.group(1), c, cls, text.strip()[:120], sheet,
                            ';'.join(mb)))
    w = csv.writer(sys.stdout, delimiter='\t')
    w.writerow(['sample', 'chrom', 'class', 'source_text', 'sheet', 'manual_Mb'])
    for r in out:
        w.writerow(r)
    print(f'# {len(out)} manual annotations, '
          f'{len({r[0] for r in out})} samples', file=sys.stderr)


if __name__ == '__main__':
    main()
